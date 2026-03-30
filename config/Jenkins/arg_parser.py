import os
import copy
import html as html_escape_lib
import argparse
import shutil
import base64
import re
import subprocess
import glob
import time
import logging
import chardet
import hashlib
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from bs4 import BeautifulSoup

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Default source paths
DEFAULT_SOURCE_PATHS = [
    r"C:\JS\ws\develop\sw\Release",
    os.path.join(os.getcwd(), "images"),
    os.path.join(os.getcwd(),  r"BVTRBS\03_VariantDependent\Customer\CVADAS_RBS_TRSC")
]

logging.info(f"Final DEFAULT_SOURCE_PATHS: {DEFAULT_SOURCE_PATHS}")

# Default destination directory (current working directory)
search_pattern = r"C:\JS\workspace\**\BVTRBS\03_VariantDependent\Customer\**"
matching_paths = glob.glob(search_pattern, recursive=True)

print(matching_paths)  # Print to verify what paths are found

# Ensure we handle the case where multiple matches or no matches exist
if matching_paths:
    DEFAULT_DESTINATION_PATH = matching_paths[0]  # Take the first match
else:
    DEFAULT_DESTINATION_PATH = os.getcwd()  # Fallback to current directory if nothing is found

print(DEFAULT_DESTINATION_PATH)  # Print the selected destination path

# Create report paths
EXCEL_REPORT_PATH = os.path.join(DEFAULT_DESTINATION_PATH, "Software_Test_Report.xlsx")
HTML_REPORT_PATH = os.path.join(DEFAULT_DESTINATION_PATH, "Software_Test_Report.html")

# Embedded logo images (base64-encoded PNG); sidebar logo is scaled to 285 px
# wide (matching --sw) to keep HTML size small; header icon is 36×36 px.
_MAGNA_LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAR0AAAEXCAYAAACQ8ZAdAAA19ElEQVR42u29Z5Bd15Xf+1t7n9u3bzcymJPEHIYkSIJBFEU1KYJiEDPRIAKhVzWvnj1VY7vK5fEHz3s2zSrbz6+mxh8cys9+VXK9YQBJSJREK1CjYYASNRKjIiWBpCIlJoBE6HTP2csfzrndFxD6nn2BDjesVQVJpBpA97p7//9rrb3+awlmZkdoCvLsyIi/ZseOIBAa//79C05ZOTAweN540PNBL1A4w6HHKDIkUFN0CJEAul9gQmBvUHnTOX1VkB8mmfx4ePnUz2THryam/65RPADbCQJq3u8+E3OB2ZGADaM42U7W+Hd71552QRrkeuBahItAThxyToQcjQJK0BwtAoogSHEQnYBHEIFMlYmgUwJvOPh7J/LUeHB/d/RLP3vzAAAy8DHQMes/sNm95kMrkqRysyJbFb1umfeVAjSYUkWVAAXCtD53Kopq/hXiwFWdMCg5NO3Nsl3AlwL68PLTX/u7xt+vo/hm4DMz0DHrJcBpuuBvrz3l+GGqf5Yqfzrs5SQF9mWBoJqCCIKTIztjqqCoqgIDzvlhJ0ypkiovpIT/d29l8qFTvvvbcc3/HmlO78wMdMy6PLrZDm4DZLvWnrbcB/mnOPnzZd4ftS8LTAXNihN1pEDT6ntQUc1UxA05cQPiGA/hJ6r675a+uPMhi3oMdMx6MLrZvfaM9Qny74edO31fCNRDSBHxssBnSSGoqg555wdEmAjh6Tr6z1e+8NqLxf/vLOox0DHrRsAZIZEdpL9ce8rxq6j+VdXJllSViWxxwOZQ4IOiSxPnp4JOpKr/90tLd/67a3eQWtRjoGPWZWfjmZERf+2OHekfLjpz3RLPZ4acnLwrDUFypHEdBY5K5gS/IvHsy7Jv7g3p1hNe+uWvnhkhuXYHqX2cBjpmnRzdgGwfxW3YTvb+Jaf/g0TcfxFIJkJIRSTp4O9bUc2WJT6ZDOF3Y5ptOubFN77ZiNbsk+0Mc+YCs9kAZ9clp//7pd7/t7qqHw8aOhlwCgYVEUn2pCEV5MQhSb7+9sWnbZQdpDoyktina5GOWQcCTqP/ZvfFZ/znFRX/57vrWabz+CI1XxaUUBHcgBP2ZNmfHvvS6/9DR0YS2bHDIh6LdMw6xZ4dGfGyneydi8746xUV9+e761mqgpcuJCcnuBTCZNCwzPvP5BHPjvSZESzisUjHrCOinCIKePeS0/9yVeL/7fv1LEXEd/sZUQgJkEDYh95y7Iuvfc1etQx0zBb7YhaX8J2Lzti8ouIe2peFNNCdEc5sqVbViQM+mMz0o0e9vPMn1sdj6ZXZYgHOfXkN571LTzu/5vnvY5mGrIcAp5FqTQbNBkSWJ8IjeuGFw9xX1LDMDHTMFjT1EO5H9MYzqmTyNxVxw3VVdT14GZ3g92ZZurziLnjH7f9ruZ/AqJ1/Ax2zhbXRUSeQvfuW/otVleTivVmWOinm1fTiYRdJdqchW5b4f/juRWdcL9vJpufzmC2YWXjZv1GOA/TdSz98VjVUXla0ktJ9T+PtWoAw7ETGQ3h15R65mC0769yP2kwei3TM5tvuAwGVkPw/w14G63mBQ/rgwLuxLGQrvD9313L9R3I/4dmREYt2LNIxm9cop3it2r329GsHxT09HkIG0k8XL1RESIPuTrJw3pIfvP5OcRks2rFIx2xe7Lz8cmXKfV4E1b7jHjcVQliRuNVTifsnAopFOxbpmM1vlPPu2jOvGITnpvKhfH1HPgphQESmgr4lLpyz6oXXP9B88qBFOxbpmM3LBx/Cn9WcCKp92SAn4CaDhhWJO05w6yGXgdjJsEjHbA7tPnD3Q9h7/qnH1ivu597JslRR6ddzoGTD3rm9IXx39Ys7r8JmLFukYza3ds3IiAOYTPwtyxO/rB40kz4mHhXc/hCowGW71px+nkBQuxMGOmZzCDrX7AgA3untCip9HugKiKpmSxOX4OQWAApgNrP0yuyIM4m8SPr+BaeszJLKzwecO2pKVaXPz4Aq2dLE+X1Z9tSqF19bZ0JQi3TM5soKnVE9SS4Z8u6oKdUgRjoguImgqMolH1x2zuoixTK/GOiYHbG9PSIAXuSKqgj9+mp1qBSrrhpqXlZOTdUvaAZoMwMdsyOxY3ZokWZdkqONGJvP5FhhUBzi5aJmgDYz0DE7AkKX7XnaIMpZdVUQSyGaoh0UxQnnNQO0mYGO2eESefFf76/50HLg6NSKFgf5RySP/vRDAGy3rmQDHbO5IHPQ6krgqFTVpuYd6B3JVBHlOLUGQQMds7mzeiVb6Z0kSn+MsWjH0jy2Wf6HC48dKqJD84+Bjtlh22h+gZJMlnok34RpdkCso3maVasNra6aOwx0zObIAlIp9pAb6BxkOehQGR+bsL1YBjpmZmYGOmbd+WE7JlWtXnHI/Cr/NVUfGrS1wwY6ZkdsjSfghD0pakXkQ2RXDlB0DPaMmzsMdMzmis0n3QdZ0HreDGd1nWbzeYP2npO/+9uJIvIx/xjomB0JkwOsSOq7gfcSEbtUB3pHfY7EbwmozdQx0DE70ggnv0jCC6/vAX5fETHEORB1tJhTuhOYbjEwM9AxOxIbxQmowE8rDbGR2QGhoDj5EWCCTwMdszmx4iKp48Vc66kGOtOhoPjxoCjhB4AJPg10zObEduQXSdC/HwsBFbHNB3mEowMiMqnhHYE80tlu2isDHbO5sACQ7au8MhnCHwZFxF6wACUMOkHg+8XuK2dFdgMds7nIIEB1FH/0z362V5BvDjpRUTLzTF5EDkH/FrDB7AY6ZnNqb+evMkp4XHORY38PZQd1In5Plk0mIl8u0lBLrQx0zObMduSRjU/Tr+1Js3cGnPi+TrGUMOwcmcq3V7z02k7bBGGgYzZPKdaKH/56t4g8PuwEVPs6xfKCiPAZgGcttTLQMZt7294AIMd/2x9C5kT68gwohEEn7oM0/Lo+MfBFBblmxw6rcRnomM21bdhOpvfhVj6/86XJoF9d6p1Dte+U1aqqQ05E4b8c+5Of7Ht2ZMTbq9WCRdxmfcjyTiC8f/Fpl3rn/r6u02dB+uTnD1URqQf97S4XLjgtl4eYHs0iHbN5ZJqgo6N+xUuvPz8R9G9WeOe0j2o7jSgng399+guvf2BRjkU6ZgsU7QA6tubME9JEfySwrK5Ir8/aUSVb4p3fl2XfXfXSa1dxH8j9jWmlZhbpmM1rtMMobviVX/xuKuP/XOq86/WXLCUfYVHXkAXhHwmE7T8pdu2ZWaRjtiCXUBjFue1k715y+tdXeL/u/TSkIvTkcHJVTVdVkuTdqezfHP3yzn+po3jZbl3ZFumYLSTjKOehAaTm+NP9IbxddeJVe69BTlXT5YlP3k/THUctP/F+HcWbsNMiHbPFupAF4//holOvX+aTJ+uqmuYvXD1xPoIShr24NPDbCXVXHP3Sz9607mOLdMwWk3m2kz0zQnLcy298fSyk/8eQd94rWS9IJIISqk5cpnywpx5uP/qln735GHgDHAMds0W2a3eQ6shIctRLb3zm/Xr2F0sTlzi6G3gagCMwsVezu0784Wsv6ih+A1bHsfTKrHNSrZGRRHbsSHddcvpfLvf+3+7JQgj5QekqggpKVssFrWNjWbb+mJdf/6qOkMgObK+VgY5Z5wFPfjl3XXzmPx50/McMmMx1Wl0xbVBV0yXeJ6nq2x9kuuGEl3fueGaE5FoDHEuvzDqUiXaQPjMykqx66Rf/aSILd4C+v9R7H1TTTs61Qt5snK5MfFLX8IPdU/WPn/Dyzh1qgGORjll3RTy/v/C085dV/P8/5Nwlu7NMVQkidFTUE1SzAef8sBP2hfDYrrT+D0995VfvWy+OgY5ZtwFPcWmfX3v80Jm65D4n+s8GRPzeLGQiIotd61ElQ3ArvZOxoG/Xg/6LlS/t/AzMCFvtUzTQMes24Gm6vG+tPe2qGu6vhp27clKVsSxkgojIgoKPFg2MsjRxrh4UVR6cdOEvVz//2m8aujITcXamWU3HLIaZgoLoKP7YF17/9l+/sPNjezT870H1pysT74e8cwpBVef7iT3kfYvIssT5Ie/cRNBnpoKsW/LiL7aufv613+ho3oNjgGORjlkPRj2/W3v80FKGNwF/VhF3aUVgf1DSoBmokk8lPGzluuZ7SEPjz6qKuCHn2JtlmUO+nPnwX1d8/7UnAR4D/2PQ+y2dMtAx681zo6O4RoH2PnD/5OLTPzng3J+q6idq3q32wKQqk0HJIKAaGgdOEZE/BhgaW0cFUBFfEZGqCBUHEwFSDW9kga/g9DMrX3jtxeL3yfZR3AYrFhvomPVF1COMjjrZvn36wu+94tRj09Rf7ZR1wBWZcnbNS61awIyStwMfvNTYCdPPYQqMBaWuutsJP/bItxCe2je157njf/DW/kbExShiL1MGOmZ9Cz55fbAZBBRk99rTTna4s1E9K1POFOQEYDXoECLa9IfsUfRdnPzaBX7hXfj5vsz9/ISXd75zwN+Vq8PVXqXMzMymgUZH8To66ufqz3tmZCTRUbwaSVqkY2ZWGgHlUZDw9kh+1o7ZoWzPX5c0LzI3npnk2ZERdw3wLHBN09eZJ83MzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMx6wKxPp9w/nrxXpFd81fhZAovb1euKX9qD5zDD+ovMzMzMLNLpZHOADlUqF4ZU/1UREfTK7KHgwGVO//tkln2tiOQWUjTpgWzQuY1O3ajma258D50dVfV/McnkL4szYxqxgywxF8wKOmlIwz9wInf1JNsEWQ58bRHSgBzAVf4vhD+RHmS+INnLBP6NgY5FOu34RIGlgy55VeDY4uD0mq/qIaTnT8LrC3g5PJDVSC7H8VwRYfWSXwPgVfVHE5qtLf7ZajsW6cReDH+DwAmaX4xKj/2MqUBNXHI3If2rBQQdARDRe0Gc5n9nr53BICJrBjW5fIL0uUVIX7sijTA7KCcHVEU29zBLuWJS3+biDCzEpRAgBYZUuENnAL7XLOTIqvdaNmGgE+uPbJDBk0X4ZDGawfXoz6mCrKmRXF6A63wDgAek5v0Ngpzcg6nV9M+Zz1yVO4FlBdAa8BjotPaHuPR2geEePzCZgOgMIy9IBEmQjU3/3IsmQObg+EHnbmSm18vMQOfQF7G4D5u090NjVzDyHeQAO5+RhwDZMByrwg09nFodCLDqGim6vWAZ6MyecgxRWQNyxQKlHIv98wYHJ9a8v36eGdkDkrrkTgfLezi1ak6xRIR1g3AyvdXnZaAz175QCeulf14cAoAG2TLNzvP396hDN/bJWZIifR0Wl9xhd81Ap1VqVUFYr/3jm6LoyfVDcFzhAzcPZywsZeBskCv7ILU6IMcC3XRg6m5moDNzCXQYPwJyTh+Fw42i53Jc5bZ5OhMOIJVsi8AAeXG+b84UyGXDVC4oMMjumznhoFBH2CxNaUe/WN6YNM3Ic/mzS1MEuUH778xlAkmQsN7um4HOoS7GKkRu66fw/wBGFj46AGfPcZTnAK3irwKZ6z+7K+6XAircQx7lWYploDN96WTQVW50sJref1mZjZEHnEs2zse5EGFTP0aQhR+DIGdX8VfT+y+iBjqRVojydBPz+4LT8YwMbJxDRm5EkEsRuUX797wFAZyw1a6agc40Ew0y+CGB67R/u0cLRuacwST5yBwxsgfIZQ+cQP/2qvg8xZJPLYeVfRpJG+j80c/v0lGBGv2tk8mFitmc9dI0eoC2Sn935TZeCI+adO6GZkA20OlPy/tStC9kD3EplsjtwNIjZGQBQg1OEmGdmv6oWRYBfS6LcH3+s+sQlTUichFW5HMFI59Q8/5IhYr573PJnQJDmNK6IYu4fhBOoTeHwhnoxP7sKmGTLNxMma5gZNUjlkXkTK660SLI6Z8/FRj0zt3d7ymW6+dDANQQGdX+9sUfMzKsq1E7icOTRXggVGAN0hfC2egzp0CGbKHPSc718c8tQ95fJ/BhTAV8MCMP4+p3HuYZEYBE/BaxUZ0Hg7EKcvEQlb5O5/v5oqkGafTm2LyTgxiZGUV4O6DRiCAHEO60CPKPLBNwKmEjfZx29uOBECBbAkercJO9rBzyTCjIFYchVHSAVL0fATnDIsg/9k+h6r8LqNKnBfa+bNYCqDt3u7NmrVlBWcCnM4zczjlRF9jap7KHmPsWQE6ven8tvTuD20DnIAuAc+oaqZXZLIzsRNaTr9+JYeSmCFJu6kPhbPT5E8AF+vb8uT78ecMQHKPCFdqnTBPPyJw1jL8qkpEdQOb9RwSOos97UVpF2g1ZxNI+FRj324ULgIzBO6L6ClZELmXkIPK/RTJy7scse0FhN9O1IbNDRYQOVmczg9P6KiLs15pOhsgDYkwcwcjcGsnICvhxeFOUJ4svtOfyFs5S1b6URfRrTQcJA19U2Mv0WEmz2Ri57io3RzKyAILTbdaJXHrvFOHqATiTPnvl69tC8hhjv1f4WqGANkZuScrRjJwBOp5lfwf8sgApS19nAXSBqnPJhn67i33dkSyiD02zs9kh/VTIIq6twqkRjKxAAoyr6uft2bzUtwAbCp9l/XT5+tEyQIey7KkAbxojt2TkVKAqLomVReT9byrb1HRXpSmWwIW1JLmEPvJVv4KOAsl7sBfVzxkjxzCybmRmBKmUALqMk76g6EtNv8fsEL4SQDPd0m9o28e1CkDlEWtki2FkubRGspY4WUQROcqj0uxrs0MDushdzP8+eQOdDrAAyATp9xT9QfFhW7QzOyMLohua0q4y36LBP6YwXtQsDHgOff+Cg5Nq3n+CPmlW7fdIxwMpyCOWYkUx8ihxs6QD4CaZ/KUq38CaMMsAWjXop/sFmJ194BBC+jmFKWPk1ows8OGBGUb2MWfLizxoTZitU9F8lKm7YR73yRvodBjouCn4uSrfEGPkUkb2MzOIysA5A9gf6l8O8C7WhDmbNXp2lqpLbu+He2lix+lZybINY+RSRlbhU0tYcjSRsghgt6p+2WQR5bl+G02YBjpdbhlAJdSfCLlQ0Ri5BSM7WFF3k59qAFHkKXvApgiW3kMFuXIJA+fS47IIOwQFI++Dd0X5ijFyubNE9d5IRs4Amcyyb4HuZGZkhtmhU6xK6kLPyyIMdJo+eHH6kDFyOSOL8LEq1dhxpB6YRHnMXghb+7ZIsXpeFmGXa+Yi6FiWPQO8boxcyshVcVksI4f8RrlHNb9I1oQ5+10Mgpw3SHIVPSyL6GXQSQZd8hcFa8RkDQkwgerjxsjljCzopkhGDoDsp/5DVf0+Joto6SsBnMjmXkfXnvyZBhg4S+D+AQZObYeRRd02zf+3FZRbpFgg5w8mSewyvXznU96zY9bCTwoE0dtWwTJ6VBbRs6DjXLhdkCHnQqw6uhhlWn+pSaho0c6hLROATDe1kb5CqHxRYb8Bemn6etyYczc1AbaBTqdfiDzs141FD9tm4gtzxWWQh42Ry1OsYn/T0ghGDoAbZ/y3Cn9rTZitU/3cP+7T9Gia32ug4wEdTJIrBblQIRNkzSDJZZFpQMHI6WcLoaIx8uznJnPI8YM5I8fKIkREH8aaMMtSLBH4RLVHV173ZCHZZbqpONX1vDDXVhrgJuDXwDM2yrQ1IwPq1G0hLnLJAK1l2d8G+D02OK1VipUKDLr4wWkGOouZDwNLg8gdRXhSyQtzcndkGjDNyEp4ABtlWsrIKnyiBicRN8rU74I9qH7RXghbn+Vc3KZbGlGlgU7nplZS8/4GB8cz0xMSHJxQ835dZBqQAToRwpMB3rEUq5SRl9AuI+ejTHs20p6rMoEglwxRuYge69nppQ+9mEsimzlQBd3491uIU0c3PuD3UX3CZBHljEzesxPTf9MYnPYdRX+MNWG2JD8BUQmbmkDeQKfDWDfU4CQRPqkHRjReQRCub5pXEvUBquNBY+RyRga5fJjKBZSPMp0enOaQ7ZZitb6bxQvhemCQuH3yBjoLfPjBJXdIPmu2+QNqqKOXqXN3HPD1JYw8mWXfAX5ujFzKyD6VcE/kmQoAaXCPKtQxWUSruxkE+fCA99fSQ6NMewV0irB+eh6JHIpiQe4lLl1qMPJUUH3UGLmckZ3IBmCAyFGmU0y9CnxbTBZRCtCe6ZKBpVedFOIPU7kQ5PJZdi010oArhqmcTxs9O0HddhMqljMycMYw/mOR58rlDpZt5r7WZzvfec7NS+AoekQW0QugIwBBwj2SA0PWIg1IgoR2Nhq4OvUfovocxsgtfSX5f2xtJzJNQv0LAd7HXghbne3MwargKrdFlgYMdBbgQ0mBKsKGkqJvozDXmFcSU5jL/yyRbdasE8HIIrcsh5VEjjLdD2+L8jV7ISzN9TXodINr16f5rhe+/yr+apCyoVJFGiBnV+PTgIYs4gsKe42RSxn5qEnnbiauH0rIB6fZC2F5+QARrh5g4Ex6QBbREx+0CFsiRYQhP+nRaUAA/Di8qTotVDRGbn2k7qUNWcRYlj0F/AZ7IWzpK4Gqc+GuXri33fzNC5Atg1WI3Kpx7OqLpXG3LocVxBXmGozcECoaI7dIXwVGIoWKjcFp42qD06J8WzRh+m4nvm6+QB5gyrmbHKxuA0AyB0dPOncjbcgixnOh4u+MkVv6NhWoiUtGI89XzgEqD2kPj+eco3uqxcSE2MFpBjrzYPmkP9xWbb/Ooqj7dGQa0GDkfSZULAeegpHvIU6omAEyTvqCoq9gg9PKUiwQ3dgLCNqt33eowqnANW2mPS5f48q1VaqnEVeYK+6SPGyMXBp9qiAX10guoVwW0fg9oWlwmoFOqxQrH5w2TBf37HQz6CAuWS9QpT1dyvS8EnFZbBqQkQsVvwv6I2PkUkZ2TYwc9UKowX9OYRLbJ9+SaB1yYs376+ji+mK3gk7RIaybdQZIDicN2ED8vBKfXygTKkYy8t1AjUhZxCSTrwPPYqNMywBaNcjWbgbmbgQdD2iN5GJB1nB46U5zGnApbcgi0iR9RGHKGLk1Iwvy4Sp+hPhRpijhIbGhaS3PbWNiwjAcW5Cl68YD0m0meXlFtxyhWLCYVzItAo2SRUxN8QvgW8bI5YzsJHqGUQYwEcJXAuzCmjBbnf3MwXKdkUUY6CyA01OgJiJ360zUcthpgIjeAQzRhiwiGCNHMbIKNy+DVUTKIoD3VPV/WhNmaytGmXatLMJ14fcrQ95fJ3By4XA5gj8rCJxc8/564nt2qITwhAkVoxh51aRz7QgVxTseVmvCLDu3CnxsCQPn0oWyiG77YHOlf15ImwuUL7RV0/NKYnp2/D54V5SvmlCx3FlOXexnFQDGsuwbwE6sCbMloAtUUhdGu/Eeuy5zdlgCR2s+kvRIUqvmNAAVbihGmUZHTqL6gAkVIxhZuKoKZWLc5hRrAtXP2gtheWkA1XuIn5hgoHM4AAFQd5U7XLxuKjYNWJ655PbINCBnZLJngV8ZI5cyctW7ZEPkeStkEe7Rpn3yZrOVBkTOGyS5ii5LR7sJdEL+DU/35sxtGgAbI9OAaaGiMXIcIytsIm61c0a+T/4VVf0eNjit5X3IlcjaeCG09Go+kH0pA2cDH2VupQiNYvBVAwycQxuyCFG3zWQRESkWnD9Icnnk55b3UIk8Ik2+NputNCC3roJldJEsoptAh7oLGyUf/j3X7JcJVFx8Ya7ByC+p6ovGyKW+BZl9aP6hIlpC+jmFfVgTZln6etxYPjgtpjRgoNPOwc0Pn47OU/G2eV5JhXhZRECmhYp2MVr5VuROYAmRsohx+K0qT2NNmGXIo6jrqp6dbgAdD+ggyUcF+RPmpy+hMa/k3MEk+QhtyCI0pI8rTBgjt/RtVqx2vj7y3DlAEGvCjAB0EWHdIJxCl/TsdE/FW2TzPBdtMwFcNp0GlFkhVOSXwNPWRdvSDu6vipFF6HAIXw3wh+mo0uxQKVYqMCQu6ZpRpq4LnJoBy4LobXPUmzNrRKVAyNOA5cQV5hwgSmiMMjVWnt23gnB9DU4gsmfnPdjbtE/eQGeWO1LIIjbSJbXFTgcdD0jN+xscHM/8VugbPTvH1rz/JG2MMi2Eiu9gsogy3y7BJXe2dfZUbFtERPlBkMuGqFxM3OA0A52IsLwhU9AF+vtiNxo0aj+7RXnCZBHlzgLdEsnIgXxw2nOgP8WaMMtKA07j98kb6LRixxqcJML1kdse5iQNOKgwFxVZBafbjJHLGRnk8qbVzjGyiFQQ2ydfco+LwWnrgUE6vGfHdfghBZfcKflM2IXQlxyqMBeTYslkln0T9OfGyKWM7FMJbY0ydcE9pvnnn5gLZ73HQZDTqvhrOv1udzLoFOH39NyQhUJuaZpXEluY88BUUEwWEcHITmQ9eT9UVM/OPqZ+CvodeyEsBWh1wmY6vK7YqaDjAR2qVNaAXL7AUoNGYe7SoUrlItro2QnqHtHp+c1mszEycNaw9x8nTqhYDE7TB7HXwdLSgIp8amn8HjgDnYPSHDQN98jibDTMC3PpdGEuapRpnfoPQb+LySJa+kqAEGRTJCNnAEk+OG0P9kLY6s5kDlbVO1wW4TrUeSlQRRhdpOJsozA3Sl6Yix5lCiZUjGBkVLiVuBElCvj98JYoT1qK1dryOTturobc9Q3oOECq+I+DxAx/msc0QE7Lv4+oNKAQKlYeN6FiFCMfM9guIzt90EaZRpQGhI9X4TQ6VBbRqR+eyvzLHmLSAEWInVdSCBXH31TlKWPkGFJ2myMZOQCMZ9nTCr/FXgjLSgPVNvbJ9z3oCJAthdUIt86z7CEmDRAR+VTkRoPpKE2c/g3GyGXpqwhcV6X6YeJHme63wWnld6howtxI/CLJvgadxkjSm9ziV+AbacDq1FVuiUwDMkDHs+xrAX5vjNzSt43Vzm0JFUXlERucFpFiIRe2sUiyr0En5CdS79UOqYcoqKKxhblmRv6CMXIUI2+KZOQMkHHS51X1FWyffFmK5TQfZQod9nTuOux7CVWqpwHXdMj31/j7Pz4AZ9FOYc7Lg8bIUYy8dojofqi8fWJmlKmBzuzpK5JPTIhdJNm3oIO47G6BKp3R3NQYCTngXLI+0mcZIBNp+j3QnxgjlzKy6IwsImqUqQb/WYVJrGenJYEXiyTX0WH1xU4CnbyTV6e3PXQKMhdpAI0dQ7GyiFQwRo5hZERGgRqRsohJJl8DnjXflgK0htB5sohOAR0PaI3kEhFZQ2elJb5gjQvbHWXqg9uuUMeEimWM/OEh/EiTv0vPrCLbMFlEy3NbvL7eOAzHFmTpOuVD74hoAkBFN0tnSgiCAGTaThrg9jL1M5RvWc9OOSOryL3FP8eMMmUi1L8U4D1LsVqXBhwsD+0OTusD0GnIHoYQWd+hM2l8kQbcQdxGg+mfQSQ8ZIxczsgINy+Bo4iURQDvifJlG5zW2pp6djomFXUd8j3IkPefcHBSJ4WBh2CNE5sKczE9O0gIXwrwgTFya98KrKy7yq2RKVaO504f7rCIvRPTVwX56AAD59IhsohO+bAamwK0gy9mY5RpQxYR1bNTCBW/YilWubMEvTeSkTOAsSx7VtHXsCbMMkCviGQdM8q0E/pgsmE4BuGGBRpJeqRpwPVFYS6GNfLllk4fNqFia98WzvrYAANnRvrWA5Moj9srVus7lvfsMEr862vPgw7qKrdL/NqXxU6xlof4HUMNRn4a+JUxcktLBQYSFzZE+raxT/5hG5xWeseCIOdVvf8YHfAyvNigE4rc6t5uKXa0WZjTgl3GTKgYx8iK3sOMLEJKAF3GqL+i6AvY4LSW90wAH9zGTkHBRUXgJQycg8iVdMG+HmY2Gly5hIHzItOAAqfkUc2/3hi5xXkAOX+Q5IrI89CQUmyzwWmlpQGC6O3AssXOKBYbdKhLtlHyId0dvTajmWEFKplkm9pIsWSc9AVFXzJGLmVkQdpa7YyG5HMK+7HBaS1LAwLH1by/oQmw+w50MiBxwoYu2xfVGGW6gZmNBjERUgB5zBg5xrfT/VBlRBQAN8HEb1R5JjLl7Wfkaby+LqqfFuuie0AHST4K0jH9A22mAWdV8VdFskZDqLhdYdwYubVvBU4a8P464l4zHSCoPCAzrG52aEAXEdbV4MTFvHOLetGd6JYuLa4GAZzI1jbSADfJ5BsKO6xnp9S36pleJR3Ts6ND1J8M8DbTDXFms6RYw8S/vvYM6AiQrYalQeS2RR5JetiRWrHR4JaV8U/9OSMzLYswRp7dt6LKTUNwHOWrnRXwu2BP0+A0A/QWzipeXxdt5MpigI4HZL9zN7n8UGVdeAGnNxqM5xsNYmUROhHClwO8g8kiyny7FFe5LTJ9Lb5KHrZ98uVlDZArhqlcsFgp1mJ8OAFQ1DXkBN168Yq2Ercp8udoNGXtFlUTKsY4V0OsLCKQD077DuirWBNmS/IT8EHCpsXCgIX+CwUINTgJ4boOlz1EpQEiXDcIH2onYnOOh4yRIxhZ5MoKlfOJ3xZRt8Fp5cevaZHkAIswytQtwmECl9ztYJgOm916GACaCgz5mVGmUcrz/Vn2DWCnMXIpIydOQuz+pgBQD/JZzc+VNWHOfueDIKcXiyQXHAcWGnTyp2N0Y4eNJD1s4MmfoXQz8TuGEmAqqG43Ri5nZCfc08TIZWfL1an/GNXviDVhlt5DJ2xdlA92gaOcMETlIkEuozc2JTTa8C+uJUnszxQAgrptTUJFKyjPwsggZ1e9v4r4nh0Qebgp7TI7dGkAFbmljUWSXQk6xUjSsFEaq0R6Jw0QzXQTcdFbAKRO/UeKfh/bFtHSVwK4+OHixeC09IkAe7EmzFZ3MXOwaipfJLmgtVW3gD9kClQR7u6xIurh7BhqREgPWbNOFCPfDqwgcpTpGPxelCetCbO1FT07sUPpug50HCBV70dAzqC7ZA9RaYDAKTWidwzlH3BIP18IFS3Fas3IRw/ibopk5Lzx0ganxZxbFfh4seBywe7kQn4gSmBzjxZP894jkU2R4BEANw6/U+XrxsjlhCzi2pJFjGfZ14HfYC+ELQFdYNC7bP1C4oFbqB9uKawWkVu6VPYQkwaICjc17RiKkkWI6oMYI5emWAjXDsLJxPfs7FfbJ196N3MU1w3Ev752Beh4gMxVbnGwmu6UPcSmAcs1L8zFAGvOyGRfD/AHTKjYyrepwLC0ub9JVB62ffKld1MFuXiIysUskK8WAnSKkaS6SXv8UuVaiLCVOHZtfMB7UP28ySLKGblJqFjmp8bgtO8p+gPshbClrwSczsgi5j0gcAvw54cqnIFwzULmjYtgjR1DVw0wcA7tFOa8PGKyiHJGBrm8kgsVY0eZNg9OM9CZ5dwW6etdwCALoBJYCNBBXHK3QLVHU6sDUiyBxLno1v2MGaHij7GiZxkj+0TCxkjfNganPaowifXstAwMQE4d8P5aFqC+ON+g0+i43dgnTJ737KhuZGbucwyLp2KMHMvIG8gJrIyRG4PTXgO+Zelra4AuBqd9mgWY/DCfIOABrSXJZQIX0R3bHuaGNUTOGyT5CG3IItLgHlOoY0XPMkY+fShn5Jjzm+9VIzywUPWKbk1fG4PThuEYygendXSkg2a6sc9YJuRrPXVL7NcDboqpV1G+ZULFckbWIPe2EWkzEcL/DLALa8JsWRooXl/bG5zWQaDTuDjDiNzdZ0VS39ho0OYoU0Rkm53/CEYWbloa137RiDR3iartky8LEGhrcFrHgY4DpOb9dQ5OordkD7Gscey4czcSP8oUCfUnFD4wRi717aop526jDaGiOB4wWUR5OQSRj7b9+tpB6ZUSWHAxWSeRBhrduq+A3w9vYULFOPTRtmQRjOWD097AXghb+kqg0sbra8eAjgOyITgOkRu7fCTpEaUBIlzX5o4hyZw2tkUYI89yvnJVP1dXIVY8nAATNjgt0rfoJuJfXzsGdMBVbpMO2Ju8mGlAsWPo7khfB0CnsuwphV8bI7f0bSpQFZfEMrLC9OA0k0W0vrsKcm7TPnk/PwAxt1aMJA339ntRQgHynh0iWEMLRh7TGVmEgU4LRga9hxmhopSkWK5O/QeKPo+9EJalWO28vi466DggDDBwLsiV9EdvzqwpFnlh7vJhKhdG+iJ/+FJ5SHPAMUZuwciCXFgjuTzSt42B5A/bPvnS0gCI3EbcPvmOAB1EsnskZ+1+TK0OZo12dgw1hIovMCNUNEae3beiMh1JxkXgIfm8whgmi2hZGnBwQs372NfXRQWdDKg4kXtMwNiUBhy4YygmQgqKbDNGjvGtrD86jpED4CaY+JUoT9Gfr6rtVQbyJsw599NcgoIHdJDkSmBe3/m7LA0I5DuGRpr8FMHIfrvChDFyWbrEifvwnyDuxc8BEiRsE5NElKZYh/H6uiiRDiLca0XQA0FEACcSu2OoIVR8Q5VvmC9LfauIxK6ozgCthvDVAG9hTZgtUyyBJbQ5OG0hQUeAbBUsU+G2Hh1JekSsgXDLEjiaNkaZItNCRbPZfSsINxZjYsuEigr4D+B9VJ8w5XlEjjUzOG3OiG+uQMcDjLnKjRI/I7jfWGNlyFv3YwA5A3QghK8EeM8YudS3y9RVbo/0bX6VHA9b3bG8XALykaH89XXOUqy5cniBgrpV7HLMes6Duk3EpUsK+D2wS5QvmSyi3FlK2BLp2wyQySz7NuirWBNmS18JeJWwYS7xws3RnxEG4WSBT5iobvY0QISrB+CsdljDqT5oPo1i5CsH4NxI33qgjvJZq5mVHL88nlzPzOurzAVgzMmfUYwkHWIBZqx2qaUCA4lLYncMBUD2k30TeM0YuZSRKyL+njZ8S6ruERucVh5QgJxV9f7qucKMuQCdDBBFN+lMnm02C2uEvHU/Zpd7Q/cyGVRtlGmEb53IKHFCxUAui/gx6N/b4LTWvsr3ycvmuUSyIw5th6isEWQtJqYr87UKckGN5FLaGGXq1T3WJIuwmtmsjMx5g0n0mNji7KsNTisvDaDCbcvj9snPO+gIgErYLHHsbWlA3rq/hbioMAAyRv0VRb+H7W8qZWQyjd3fVAxOC48r7DVAb3nHMwdHTbroffLzBjpCXr+pItxlz4/xaQAid5O37sfUvxpbGLdZ3hrl2zsifauAH4M/qPK39kJYmuor+evrEcsi3BH+Xql6f60gp2Oyh1iftSumK2QR6ecV9hsjl/r2+Jr3NxD34pdPcXC2Tz4ixRIR1g3CKUd614/UydpUYLKwvw3WkCCxo1wLoSK/Efi6MXK5b3XGt6XpLsB4ln09wO+wF8KWKZZATeZAFuGO5JtYCqtVuNlkD+2zhrYnpsuFioRthe8t02rhW4R1NTihABVXAlIJsB/VL9gLYVSOtZkjfO07XNDxAJmr3OriVoGY/TFrLMUld0R+DhmgEyE8GeD3lmK19q3LfXtXpG8LiZE8ZKNMS++8CnLpEJWLOAJfHS7oBICgulnt8B82a8D0S0tsz84eEypG+3ZzpG/zffKk31f0R9gLYUtfCTiVcE8TyC8I6DggVKmeIcLVWGp12KwB8pFK/CjT4kbJNnspLD2fCnJ5hcoFkb71QMrMC6GBziy+bdonP8hhqg8OF3QQl41K/hcb4x4+a3gXL6ZrMPK3FX6CFT3L0lfvZhg5ShYRgvuswhQ2OK1lwAFyajGU7rBe/A4HdLKcGdRGks4BaziRDUCVuLqYB1Kx/U2xvm0IFaNGmU4x9QuUb5pvSwFanRD7QnjEoNMYSXq5IGvo720Pc8QanDns/ccjfRkAMvUmVIzz7dnD3n+M+G0RBAkPmvtKMUBU5OZlsIrDeEQ6PMAQ3WjFzLlhDQFCkE1tsIybYupV0OcwoeJc+jYDqITw5QDvYy+ELdNXB6tT525pAqJ5AZ3GAR822cPcsUYhprt1OawkfpQpgnvIehSifHt7pG8V8PvgHRucVm4KGnD3chipqGvza6Xm/TqHnITJHuaSNY6adO5W4mQRhVCx/gWFD4yRS3179KRzN7bByCJOH7bBaTHEx9VVaFsC1a5TlVz2YDuD5pQ0UHCxflXA74e3MaFirG9jR5kGQMey7BngDeyFsCWgCwy2sU++bdBxQDYExyPcpPOw9a/P0wARGKlSPS2SNQSQzOkDGCPH+Pa6Kpwa4duGLGIC1c/ZK1brM6iA5E2YSTvE1w7ooC65TWApJnuYa9ZIc9bIYlkjA3Qqy54O8Ftj5BjfRjNy7keVR00W0RrQc1/JBYMz++R9NJhEWGPbwxYrHsxPjly07t/TiCopL3o2hIrGyHG+bexviunZkXHSF1X1eeyFsCUu5LNBogenRYOOA8IwlT8BuRLrzZmXi0EupruolkyPMo0SKkouizBGLvftmlqSXEa8LCIg8qjtky9NX0HkLuKH0kWDDqmEUZnJ3Sy1mnvLR5lmujH26wtGfl7RH2BCxTLfujZ8m/sxJJ9VGMNkEa3S11AMpVtHZH0xpmCZAgMibLTenPlPAyRnjRrxsogM5GFLsebUt8XgtIlfAc+Yb0sBup3BaXHDowZJrhTkbKw3Z77TgAB8aAB/bSRrNEaZPq4wifXsxPj2mkjfOkAUedjcV5piCcL1Q3BcDFnGAIiK6BZD+4VhDQG8yNZI8AiAm4SdqjwrlmLF+PZe4keZ6kSofyXA2wboLbOhzMFyXOW2pgj8sEBHgGw5rFCR220k6YKxBgg3xbLG9GcowRg5wrcq3Lg0btplozj/PqpfMq1hSWQCKCFqXrorqRcw6dwNAsdgBeQFYw2B5eqS22NYgwOFiu8ZI5cy8qqpSEaevlDKg1bPLCU+BblqgIHzKCnDuLJ6Aeq2ih3iBWeNplGmUbKIvfCeKF82Ri53lqBbI32bATJJ9i3g51gTZhlZJiLZpjJscS3+fRiEU0T4hInfFjYNKFjjowNwLvHFexGdXpFrn1UL3wp8bICBcyJ964E6NjitNNopXgjXU7JPvhXoIC65S+Kfb83mzjKBStJe676OkT2j6E5j5BjfhrZkEWnitmnePmJ1zRYplsA51ZnBab4d0MkAUXSz5VWLxxqKbiBOTNeQRUyiPG6MHOXbeyJ9GwBXr9d/zMw+eUtfZwd0mhZwRkc6DtAhKmsEWYu12C8WawSQPxkk+UjkZxAARN0jmv9vKyi3YOTCt7FCxeKeiA1OK0lFG0PpVsLy2TKk2UAHlbBFZsSHZgtvDTFdo3W/7LwHQMaov6xMCxUt2mnByIhuacO3ECpfsH3yLa3xQnjsuHONffK+DHQasocqwt32TLj4aQDCneTjRGLEdI0tjNuMkaN8eztxQsUA+HHG31Tla2JD7MpSfUVdQxYRykDHA1LFXyNIzNAjs/lNAzKHnFDz/nriBqc1ZBGfM0aO8u2JNe8/GenbPDhy+iC2T74sxRIRrh+EDx0KQ9yhUMoJsX0MZgvAHKrTI2LLAKQQKvIbgaeNkcsZWXVaFlHm2wyglmVPFfvk7YVwdnBOBWrikjsPhTPu4HxsKaxWkZtM9tAxaYAAN9SonVgc/KhRpoHwkLFxOSMD62pwQoRvFUh25fvkbXBaRPqqeYPrH732uYNTqylXuc0d5hIts3lhjczBElz97lmi00NFOzoUwlcDvGUpVmtGdrAUl9wV6dvcj14esXpnafqqglw6VKlcxEEvhO7gw+rUenM6Mg9AtxDXI6KALxj5CZNFtAYezRPYzcT5Kd8nn6bfVfSHlmK19pWA03R6n7wcDDoOCFU4A6FlN6HZorHG2gqVC2lnXKyKCRVLUixAEblsmMoFxI8yzRzTo0wNdFqkWAgbgEGaXgibQQdxyQbJv8CYscNSLAHvZJo1YlIsmSB9TuFVY+RSRk5SCZva8C0+pNsVpsi7ms0OTZYB5NRqPjht2reuKWz0wAY9KBQy6xzWcLmYrkr8LJi6qhojR/mWu4EBIkeZ7s1V59+09LU1QEt+bjcfjEYe0MEkuUJgjaVWncwanDns/ccj04C8Z0f9oyZUjGLks9rwrcsdLI+Y+1qnooUs4lPNg9PcNKpnuslQu/NZIwTZGvv1gJti6qegz2FCxQjfsik2JQOohPoXA+zGXghblgYcrKq7yi0NIHLkLLgEkTut6NgVrHHzsviWBhMqtuVbuXU5rIxNX/fBO6J81ci6tR08OM0BUvN+nYMTMdlDN7DG6innplkjJsWSkD6hsNcYudS3R0/G+xZAxOlDGFmXEZ8CVw8wcFYDdJQg92At811DHKIuVqYSAD8Gv0f1yUIWYYzcipRzoWKsb3Usy55W+CX2QtgS0AUGnAvrAfwQHK/O/Qfyp3KLwrvjYzzZa2VbRrariUlmTR0AreCmRGRTcTHsc54tchFOSDQ8kMIHEb5NgKkKcoqIXFmUKizimTXD0mUp+v+54JJPS57HNuYgi/3q2F8OcAIDTtKtkWF9Bug42VMBfldcFPPloX0rAkO4ZEukbxuD0x4qkKlifjzkryQHdLmsir82Qbkc0detntM1lil4FVmDxssigH2i+j9EuNea2lr7VlQuIe61b3pw2iDJVwQ9D6hj7QmzkV/FCdf8LxkWIVuw9vSwAAAAAElFTkSuQmCC"
_MAGNA_LOGO_SM_B64 = "iVBORw0KGgoAAAANSUhEUgAAACQAAAAkCAYAAADhAJiYAAAFoElEQVR42rWX748dZRXHP+c8M3t/7NKfIi42qC1EpTFA26QmFa0JikgMFdmt2xIBrS3vfWuixPiWPwCC0Bf+LLGJRmKQKNcAMbxYYiRESpBAQgmCXVa63bv3PnOOL2Zme7l778zSupNMJjPznfN8z/k+58cIl3E4qICdvfHa67ck8tPofNncmqnomZ7z4Nb5Vx4rMeu1qZdL5r09u/ZuSeTpVORQ5r7ZkIbD5zYHefSdG3f9RMD8R+tf55IJ/RjwvXvT6HKyIXLlYsz6XrxbcbfFzLJtafjhuZuu+6I8gPkMYcMI+QzhAbAF3t8/FXT3e9EyFUnL9wJq7q7gYPcB8O8vycZFaNV43JUKJrivBYn0QUA+CUCnYxtH6KOdgkB4NwN1ZI33gnsAc1gAYIYNjNApzEHe7608dz6zcxMqak42sOHdBFMRFfz0hksm4Mygn/jHGwuC/GAqqDRVgucRMQHZnqYT5/rxqa2bdvzSQaXTif9PQmH4lFNwfC/p1vlXHvtPL94j+GstFZ1U1QDdhX7/Z/OLftfBToeD+Tphnc5e3lHWozc+v6O1vde8oT3VyPjM9Bl56KnFS4x+PaaZJAfcfSJfHymvIhK7MT7jM4icWt1DNwHtbSG0VsBjXqVFRHrdGJ8tvr2kIwC0SPa1NPFRZ1sTb5IcKCIl20M4eEVIfbJ4twZPumfQ9ofdQ3n0xOck96pH7m155vfic+VGXzIOZ+5m0PMhrIC72LfrlJGa51MtTV4ErikMDzpggDqc7Vr8FNBsavKywMfGYYF/LVv8LNC/uAXXFyEBvAnbgOkx5BXIFK5uEG5uw2RBZpRdLUjtbITwhYKIfhjJDKALr+P8tZDMRicZriLHL8Bbjj9fPM9G2RRAjaOXWoe2ABMudrJC2uAgCF8Hmoo+ImOkKLC4yDeATQVpWQ+hBJCWJkfamt7bNXvcYanIDB8hbSYw1VQ91LD+rx1WChsjsQpXtkK4tbgP6yHkRTuaM+wEsOzOHwvZsvHZoccXYNGdv9RJ7CZHL65TnWUKWAs+jiavAkli8apeCHuCy5OeEwojFgGIicXpGMLN4nK6AivAkli87gK8NVBox2YDaHIn0BAIUZO7V7LszwbvFAvYGNnSqMnR5Sx7wmCxQuIoMOma3DGqSOqo7HL8yEXt/D4gw/23FVJIgf0u0BPndxUSS84yX2PYngzLlZLuTtT/PkBWzOK1gXQT6vO+tuh9QA6zuFNIplV5tgabZaa7e/TODNSpD4AVIBGblTyMWSEFKuHYBfovOLw6+PHQsYrtEp8zeLMGm6ja7LBSg4QyIEGY9Yvv1PPxeC73y38uI8Jc2iqwR3Jd/Fe1WPxwWfGHJQtA1iQ5IMozQ73IBNSN/U58WzR5raIPmoBisi8i54PaPwcya1Q3UIz9y8TnSw461NmPjPDKAEf8RBded3y+zKwxi7iJ3d+j97LjLxVYG9dKvJgWSg5apiLQRuSQr5UyOIiLHAISQR6taQ+CyDdzj+VknWwicifQLDhIOetKK4RbFK4uPNcRJX9bO4RbJyz+wvNZqKo9bG+H8JVg8eSYArma1QLXtEM4WLYSXS03FeW8fG4mJxZhAadT4fkqdgneBv9bncSDrUSBbAo+4sLXfEzDK6UQ4RbgCtQfqhp9C+xXgUlHH66T2IXbN+fTRaaARE1v14qRYKDktxqazixn2e8d/lvTHtpN1W91rf+4w3KNxFtXVG8r95DjPreOvwHJxzw7Bqzg/KGuPQjyfWDJnT/VTAuO6xzg0oCdms/NrYqaMRhyE4s7jOTTqjxdMwFYsDhtIezD5YmaCeC8W7xekfCdgozV/KdJkRXBNDnWJXYMzlbIZsW08L0LWfakwbtVWGAKTe5WFWaLGpAV16ozc4jgM0Ur+Y3k3/UqsLPFtHC6Cptf/fD/AACvrN6ZddXsAAAAAElFTkSuQmCC"

def find_html_files_recursive(directory, matched_files=None):
    """Recursively finds all .html files in the given directory and logs the count."""
    
    if matched_files is None:
        matched_files = set()  # Use a set to store unique file paths

    if not os.path.exists(directory):
        logging.warning(f"Skipping non-existent path: {directory}")
        return matched_files  # Return existing found files

    try:
        logging.info(f"Searching for HTML files in: {directory}")

        for entry in os.scandir(directory):  # Efficient directory iteration
            #if entry.is_file() and entry.name.lower().endswith(".html") or entry.name.lower().endswith(".txt"):
            if entry.is_file() and entry.name.lower().endswith(".html"):
                matched_files.add(os.path.normpath(entry.path))  # Normalize path
            elif entry.is_dir():  # Recurse into subdirectory
                find_html_files_recursive(entry.path, matched_files)

    except Exception as e:
        logging.error(f"Error searching in {directory}: {e}")

    return matched_files

# Wrapper function to process multiple source paths
def find_html_files(source_paths):
    """Handles multiple source directories and returns all found HTML files."""
    
    if isinstance(source_paths, str):  
        logging.error("Error: source_paths should be a list, not a string!")
        return []

    all_matched_files = set()
    
    for path in source_paths:
        all_matched_files.update(find_html_files_recursive(path))  # Call recursive function

    logging.info(f"Total unique HTML files found: {len(all_matched_files)}")
    
    return list(all_matched_files)  # Convert set to list


def extract_summary_data(soup):
    """Extracts executed, passed, and failed test case counts from the summary table."""
    executed_count, pass_count, fail_count = 0, 0, 0

    overview_table = soup.find("table", class_="OverviewTable")
    if overview_table:
        rows = overview_table.find_all("tr")
        for row in rows:
            cells = row.find_all("td")
            if len(cells) < 2:
                continue  # skip header/blank rows that have no value cell
            row_text = [cell.get_text(strip=True) for cell in cells]
            try:
                if "Executed test cases" in row_text[0]:
                    executed_count = int(row_text[1])
                elif "Test cases passed" in row_text[0]:
                    pass_count = int(row_text[1])
                elif "Test cases failed" in row_text[0]:
                    fail_count = int(row_text[1])
            except (ValueError, IndexError):
                pass  # non-numeric or missing cell — skip silently

    return executed_count, pass_count, fail_count

def append_to_excel(test_name, pass_count, fail_count):
    """Appends test data to an Excel sheet while ensuring correct headers exist."""
    sheet_name = "Summary"

    # Ensure the Excel file exists and create it if necessary
    if not os.path.exists(EXCEL_REPORT_PATH):
        df = pd.DataFrame(columns=["Test Name", "Passed", "Failed"])
        df.to_excel(EXCEL_REPORT_PATH, index=False, sheet_name=sheet_name)

    with pd.ExcelWriter(EXCEL_REPORT_PATH, engine="openpyxl", mode="w", if_sheet_exists="new") as writer:
        book = load_workbook(EXCEL_REPORT_PATH)

        if sheet_name in book.sheetnames:
            df = pd.read_excel(EXCEL_REPORT_PATH, sheet_name=sheet_name)

            # Ensure Passed and Failed columns are numeric
            df["Passed"] = pd.to_numeric(df["Passed"], errors="coerce").fillna(0)
            df["Failed"] = pd.to_numeric(df["Failed"], errors="coerce").fillna(0)

            new_data = pd.DataFrame({"Test Name": [test_name], "Passed": [pass_count], "Failed": [fail_count]})
            df = pd.concat([df, new_data]).drop_duplicates(subset=["Test Name"], keep="last")

            df.to_excel(writer, sheet_name=sheet_name, index=False)

        else:
            df = pd.DataFrame({"Test Name": [test_name], "Passed": [pass_count], "Failed": [fail_count]})
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    logging.info(f"Data appended to {EXCEL_REPORT_PATH} under sheet '{sheet_name}'")

def generate_graph():
    """Generates a graph showing pass/fail test trends and a consolidated summary pie chart."""
    try:
        df = pd.read_excel(EXCEL_REPORT_PATH, sheet_name="Summary")
    
        # Convert columns to numeric
        df["Passed"] = pd.to_numeric(df["Passed"], errors="coerce").fillna(0)
        df["Failed"] = pd.to_numeric(df["Failed"], errors="coerce").fillna(0)

        # Check if there's valid data to plot
        if df[["Passed", "Failed"]].sum().sum() == 0:
            logging.warning("No valid test data available. Skipping graph generation.")
            return

        # Create a bar chart for individual test pass/fail trends
        plt.figure(figsize=(12, 6))
        df[['Test Name', 'Passed', 'Failed']].plot(kind='bar', x='Test Name', stacked=True, color=['green', 'red'])
        plt.title("Test Pass/Fail Trends")
        plt.ylabel("Count")
        plt.xticks(rotation=75, ha="right", fontsize=8)
        plt.tight_layout()
        plt.savefig(os.path.join(DEFAULT_DESTINATION_PATH, "Test_Results_Graph.png"))
        logging.info("Test Pass/Fail Trends graph updated and saved.")

        # Create a pie chart for the consolidated summary report
        total_passed = df['Passed'].sum()
        total_failed = df['Failed'].sum()

        plt.figure(figsize=(6, 6))
        plt.pie(
            [total_passed, total_failed], 
            labels=['Passed', 'Failed'], 
            autopct='%1.1f%%', 
            colors=['green', 'red'], 
            startangle=90, 
            wedgeprops={'edgecolor': 'black'}
        )
        plt.title("Consolidated Summary Report")
        plt.savefig(os.path.join(DEFAULT_DESTINATION_PATH, "Consolidated_Summary_Report.png"))
        logging.info("Consolidated Summary Report pie chart updated and saved.")

    except Exception as e:
        logging.error(f"Error generating graph: {e}")

def minify_html(html_content):
    """Removes excessive whitespace while preserving structure."""
    # Remove spaces between HTML tags
    html_content = re.sub(r">\s+<", "><", html_content)
    
    # Reduce multiple spaces to a single space
    html_content = re.sub(r"\s{2,}", " ", html_content)
    
    return html_content.strip()

def remove_empty_sections(html):
    """Removes empty <details> and <div> sections."""
    html = re.sub(r"<details>\s*<summary>.*?</summary>\s*<div>\s*</div>\s*</details>", "", html, flags=re.DOTALL)
    html = re.sub(r"<pre>\s*</pre>", "", html, flags=re.DOTALL)
    return html

def generate_html_report(html_files, source_paths, destination_path, keyword):
    """Generates a professional HTML report with a statistics dashboard, consolidated
    failure summary, live search, and per-report collapsible sections.

    Text and log content is embedded inline so the HTML file is portable on its own.
    Images are stored in a 'report_images/' subfolder next to the HTML file and
    referenced by relative URL, keeping the HTML file size small.  The 'report_images/'
    folder must travel with the HTML file for images to display correctly.
    """

    try:
        logging.info("Starting report generation...")

        text_content, image_tags, copied_files = copy_and_embed_files(source_paths, destination_path)

        seen_files = set()
        navigation_links = ""
        report_sections_html = ""
        all_failures_html = ""

        # Aggregate statistics across every embedded report
        total_executed = 0
        total_passed = 0
        total_failed = 0

        for index, html_file in enumerate(html_files):
            if html_file in seen_files:
                continue

            try:
                logging.info(f"Processing HTML file: {html_file}")
                with open(html_file, 'r', encoding='utf-8') as f:
                    soup = BeautifulSoup(f, 'html.parser')

                file_title = html_escape_lib.escape(
                    os.path.basename(html_file)
                    .replace("_", " ")
                    .replace(".html", "")
                    .title()
                )
                section_id = f"section-{index}"

                # Per-file statistics — guard against malformed summary rows
                try:
                    executed, passed, failed = extract_summary_data(soup)
                except Exception as e:
                    logging.warning(
                        f"Failed to extract summary data from {html_file}: {e}. "
                        "Using zero values for executed/passed/failed."
                    )
                    executed, passed, failed = 0, 0, 0
                total_executed += executed
                total_passed += passed
                total_failed += failed

                # Sidebar navigation badge
                if failed > 0:
                    nav_badge_cls = "nb-fail"
                    nav_badge_txt = f"FAIL ({failed})"
                elif passed > 0:
                    nav_badge_cls = "nb-pass"
                    nav_badge_txt = "PASS"
                else:
                    nav_badge_cls = "nb-neutral"
                    nav_badge_txt = "—"

                navigation_links += (
                    f'<li class="nav-item">'
                    f'<a class="nav-link" href="javascript:void(0);" onclick="showSection(\'{section_id}\')">'
                    f'<span class="nav-title">{file_title}</span>'
                    f'<span class="nav-badge {nav_badge_cls}">{nav_badge_txt}</span>'
                    f'</a></li>'
                )

                # Consolidated failures panel content
                failed_html = extract_failed_tests(soup, source_filename=html_file)
                if failed > 0:
                    all_failures_html += (
                        f'<div class="failure-group">'
                        f'<div class="fg-title">{file_title}'
                        f'<span class="badge badge-fail">{failed} FAILED</span>'
                        f'</div>'
                        f'<div class="fg-body">{failed_html}</div>'
                        f'</div>'
                    )

                # Per-section header badges
                sbadges = ""
                if failed > 0:
                    sbadges += f'<span class="sec-badge sec-fail">{failed} Failed</span>'
                if passed > 0:
                    sbadges += f'<span class="sec-badge sec-pass">{passed} Passed</span>'

                stats_html = extract_statistics_table(soup)
                kw_html = extract_keyword_from_tables(soup, keyword, source_filename=html_file)

                report_sections_html += (
                    f'<div class="report-card" id="{section_id}">'
                    f'<div class="card-hdr" onclick="toggleCard(\'{section_id}\')">'
                    f'<div class="card-hdr-left">'
                    f'<span class="card-icon">&#x1F4CB;</span>'
                    f'<span class="card-title">{file_title}</span>'
                    f'{sbadges}'
                    f'</div>'
                    f'<span class="chevron" id="chv-{section_id}">&#9660;</span>'
                    f'</div>'
                    f'<div class="card-body" id="body-{section_id}">'
                    f'<div class="mini-stats-row">'
                    f'<span class="ms ms-exec">&#x25CF; Executed: <strong>{executed}</strong></span>'
                    f'<span class="ms ms-pass">&#x25CF; Passed: <strong>{passed}</strong></span>'
                    f'<span class="ms ms-fail">&#x25CF; Failed: <strong>{failed}</strong></span>'
                    f'</div>'
                    f'{stats_html}'
                    f'{failed_html}'
                    f'{kw_html}'
                    f'</div>'
                    f'</div>'
                )

                seen_files.add(html_file)

            except Exception as e:
                logging.error(f"Error processing {html_file}: {e}")

        logging.info(f"Total HTML reports added: {len(seen_files)}")

        # Derived metrics — clamp fractions to [0, 1] so malformed totals
        # never produce bars/arcs outside a valid range.
        report_timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        num_reports = len(seen_files)

        if total_executed > 0:
            raw_pass = total_passed / total_executed
            raw_fail = total_failed / total_executed
            pass_frac = min(max(raw_pass, 0.0), 1.0)
            fail_frac = min(max(raw_fail, 0.0), 1.0)
            # Normalize when the two fractions sum beyond 1 (malformed counts)
            total_frac = pass_frac + fail_frac
            if total_frac > 1.0:
                pass_frac /= total_frac
                fail_frac /= total_frac
            pass_rate = round(pass_frac * 100, 1)
            fail_rate = round(fail_frac * 100, 1)
        else:
            pass_frac = 0.0
            fail_frac = 0.0
            pass_rate = 0.0
            fail_rate = 0.0

        # SVG donut chart (circle r=54, circumference ≈ 339.3)
        circ = 339.3
        pass_arc = round(pass_frac * circ, 1)
        fail_arc = round(fail_frac * circ, 1)
        fail_arc_offset = round(-pass_arc, 1)   # negative offset shifts fail arc past the pass arc

        pass_rate_icon = "&#x1F4AF;" if pass_rate == 100.0 else "&#x1F4CA;"

        consolidated_failures = (
            all_failures_html
            if all_failures_html
            else '<p class="no-failures">&#x2714; No test failures detected across all embedded reports.</p>'
        )

        files_list_html = (
            '<ul class="file-list">'
            + "".join(f'<li>&#x1F4C4; {os.path.basename(f)}</li>' for f in copied_files)
            + "</ul>"
            if copied_files
            else '<p class="muted-note">No copied files.</p>'
        )

        html_template = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Magna Electronics &ndash; Software Test Report &ndash; {report_timestamp}</title>
<link rel="icon" type="image/png" href="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAACAAAAAgCAYAAABzenr0AAAE+UlEQVR42rWWW6ic1RXHf/+199zOnJgabzXSBz05HIui0ROr9UEIhGAJ2CA5sUlUqAS84KO+KBJ8EQV98EmlaqkvPkRQUBDRFmlf2kLQSokXSKxIlRg0F3MumfN9e/kwe5Jx/L6JR+KGj5n5ZrHW/79u/y1+wnEwQfpi/aXX/SLGJ5aT34BIEd49yvIDl+z77KOBzZl82YqD7+k7Pnrt1OzqGN+OaGMJndLptsy2nEPjr4eumVoHuP8I/ysGwKP9j+R6uiNbfbwslwd/HS3K3mQIaxvicYEzh87kTj8l9YevmVkbKA9ItFLfifL/3pBYdj+8ppi4TB98MO8ggZ+9DAAx0XQ8VBFwAPd4pHEi/hhfKwIgSA46ECf/L+nTjinhlKejezFhktCHa/YdPJYz5me3B+awDfv2LbunPS0za5qi94GlTrBGcihJe7KtOAsnjj5z0AT4av3U7oXZ6c/nN0z7/Oy0H59d9+nBq6Zu68emme1/njNI3YczM6v8upkNvvXGqaa04s4+k22rY3E3mfFQn4lU7F2Y40vtpQTOBeaixXYXmj0osu+Ti6l4HuitlGAAaIWwecKijz5di95ReGzAYDI0nlwVGl5l2wphUzYLK25CS9oFFA5L3v8sHHoJCsQ2IDg0Sk+/L9wLh96Q3RJQWGInYyZBNe8c6LYtHhRceCrtp08SGEnXCspk/p8KGwfkcHgpFeuA40O+x2bA8/tF4R/n36OikgBPSn+cZ/kg8G2FcwGlwQVts5vrymBjRi8l7C+qQJ1TL6Q/ACdwXs8Lp6wg47jtOvX9DCUwIE3QWJ/MNy+l4pmOxUNApyLFpSCUyTcKlWb83fsAwg8mBuaViukF+HIQoy4DBpDMNwseAr51581x7Ey6d4niHwkO5eBphGAh6LrFrVUxraK2CL9FsHqCxtXu/kx2pOoysAVo4/6yqvtF/bz7zuEYVQAMSA0aV4B+k5vsvpOUf0vwdWb3gyYTdNtmtyQPf8rvrGKnOOiGJs2ZDMDqABCVtgsaDiBtyyq3t6YMGYnd06O33/FPRms81C/RLG0fjTsMoAQiYnumWRisaRE2mdtzNezM+8BuWgXnOfqzKtKc7QDfnjNSjgIIgLeJ14Muzw4EuKR7F1h+3+F/FewGZQiFxR1KxUsZUKiaLqErO8QNDNnY98ZRvnOIQXCQxM1AF/eXatgpD/vuRfjC8X8PgI02eJ+R7xiOaYNRASaQtvppxIMRarfNbnUvX/QhoRptMqGrW3CpY89p6HY2WgZJtwLtHFOWHagTwiaDtRm5DbMTunsJPnP8vRp2pQBTuOtkWn7F+/IbR0DkMvCriRA2DlaznVqRfeUbXZeDEfptFy4Sen4cO6Q78vJ6W2M0xIdiCWASzi8sHhCcU7FyC0FM8GArFS/0LH41VIYqhZwlpItxvVG3mh2OtFIxdQyOGKDCGlusH7ys0oc+Ir/rGBzBeTcLVJ1C3r1Qlm8lOFa3vAzOPWn2u0EPOO47xlwa8ujp102aM5g/W3eLygq5DUhyf3XM8nLcdgCuFq3LzMr/1ijecBmCuz+16OXDHYvfAN1ahZRvtlLHZfxzjEKe8FRcYabizhw8jbmkBge5tAvo4bxWsxP6eU66f5HiXw6f15QhAZNYvF0dC/tB00M3obqTBNETNya8HUzvjGG3uJiKX3YUHpH0gPdnPo70ixzf/x08hSzeFNHzJQAAAABJRU5ErkJggg==">
<style>
/* ── RESET ── */
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}

/* ── CUSTOM PROPERTIES ── */
:root{{
  --sw: 285px;
  --hdr: #231F20;
  --accent: #DA291C;
  --pass-dk: #1b5e20; --pass-md: #2e7d32; --pass-bg: #e8f5e9; --pass-bd: #43a047;
  --fail-dk: #b71c1c; --fail-md: #c62828; --fail-bg: #fdf0f0; --fail-bd: #e57373;
  --exec-dk: #0d47a1; --exec-bg: #e3f2fd; --exec-bd: #1e88e5;
  --rate-dk: #4a148c; --rate-bg: #f3e5f5; --rate-bd: #8e24aa;
  --warn-dk: #e65100; --warn-bg: #fff3e0;
  --bdr: #dde1e7; --bg: #f0f2f5; --card: #ffffff;
  --txt: #231F20; --muted: #6b7280; --sbg: #231F20;
}}

/* ── BODY LAYOUT ── */
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,"Helvetica Neue",Arial,sans-serif;
  font-size:14px;color:var(--txt);background:var(--bg);display:flex;min-height:100vh}}

/* ══════════════════════════════════════
   SIDEBAR
══════════════════════════════════════ */
.sidebar{{
  width:var(--sw);min-height:100vh;background:var(--sbg);
  position:fixed;top:0;left:0;overflow-y:auto;z-index:200;
  display:flex;flex-direction:column;
  box-shadow:3px 0 12px rgba(0,0,0,0.35)
}}
.sb-brand{{padding:0;background:#1a1212;border-bottom:3px solid var(--accent)}}
.sb-brand .brand-icon{{line-height:0;overflow:hidden}} .sb-brand .brand-icon img{{width:100%;height:auto;display:block}}
.sb-brand h2{{color:#fff;font-size:15px;font-weight:700;margin-top:8px;letter-spacing:.3px;padding:0 20px}}
.sb-brand .brand-sub{{color:rgba(255,255,255,.5);font-size:11px;margin-top:3px;padding:0 20px 16px}}

/* Sidebar quick stats */
.sb-kpi{{
  display:flex;justify-content:space-around;
  padding:14px 12px;
  background:rgba(218,41,28,.15);
  border-bottom:1px solid rgba(255,255,255,.06)
}}
.sb-kpi .kpi{{text-align:center}}
.sb-kpi .kpi-num{{font-size:22px;font-weight:800;line-height:1}}
.sb-kpi .kpi-lbl{{font-size:10px;color:rgba(255,255,255,.45);text-transform:uppercase;letter-spacing:.8px}}
.kn-exec{{color:#64b5f6}} .kn-pass{{color:#66bb6a}} .kn-fail{{color:#ef9a9a}}

.nav-section-lbl{{
  padding:14px 20px 4px;color:rgba(255,255,255,.3);
  font-size:10px;font-weight:700;letter-spacing:1.6px;text-transform:uppercase
}}
.nav-list{{list-style:none;padding:0 8px}}
.nav-item .nav-link{{
  display:flex;align-items:center;justify-content:space-between;
  color:rgba(255,255,255,.72);text-decoration:none;
  padding:7px 12px;border-radius:6px;margin-bottom:2px;font-size:12.5px;
  border-left:3px solid transparent;transition:all .15s
}}
.nav-item .nav-link:hover{{color:#fff;background:rgba(255,255,255,.09);border-left-color:var(--accent)}}
.nav-title{{flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.nav-badge{{font-size:9px;font-weight:700;padding:2px 7px;border-radius:8px;margin-left:6px;
  flex-shrink:0;text-transform:uppercase;letter-spacing:.3px}}
.nb-fail{{background:var(--fail-md);color:#fff}}
.nb-pass{{background:var(--pass-md);color:#fff}}
.nb-neutral{{background:rgba(255,255,255,.15);color:rgba(255,255,255,.55)}}

/* ══════════════════════════════════════
   MAIN WRAPPER
══════════════════════════════════════ */
.main-wrap{{margin-left:var(--sw);flex:1;display:flex;flex-direction:column;min-width:0}}

/* ── PAGE HEADER ── */
.page-hdr{{
  background:linear-gradient(135deg,#231F20 0%,#3d3030 60%,#5a2020 100%);
  padding:26px 36px;display:flex;align-items:flex-start;
  justify-content:space-between;box-shadow:0 4px 16px rgba(0,0,0,.35)
}}
.page-hdr h1{{color:#fff;font-size:24px;font-weight:700;letter-spacing:-.3px;display:flex;align-items:center;gap:12px}}
.page-hdr .hdr-sub{{color:rgba(255,255,255,.6);font-size:12px;margin-top:6px}}
.hdr-pill{{
  background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.15);
  color:rgba(255,255,255,.7);border-radius:20px;padding:6px 14px;font-size:12px;
  white-space:nowrap;margin-top:4px
}}

/* ── CONTENT AREA ── */
.content{{padding:28px 36px}}

/* ══════════════════════════════════════
   STATS DASHBOARD
══════════════════════════════════════ */
.stats-dash{{
  background:var(--card);border-radius:12px;
  box-shadow:0 2px 10px rgba(0,0,0,.08);
  padding:26px 30px;margin-bottom:24px;border:1px solid var(--bdr)
}}
.dash-title{{
  font-size:11px;font-weight:700;color:var(--muted);
  text-transform:uppercase;letter-spacing:1.3px;
  margin-bottom:20px;padding-bottom:12px;border-bottom:1px solid var(--bdr)
}}
.stats-row{{display:flex;gap:16px;align-items:stretch;flex-wrap:wrap}}

.stat-card{{
  flex:1;min-width:140px;border-radius:10px;padding:20px 20px 16px;
  border:1px solid var(--bdr);display:flex;flex-direction:column;
  position:relative;overflow:hidden
}}
.stat-card::before{{
  content:'';position:absolute;top:0;left:0;right:0;height:4px
}}
.sc-exec::before{{background:var(--exec-bd)}} .sc-exec{{background:linear-gradient(150deg,#fff 55%,#dceeff)}}
.sc-pass::before{{background:var(--pass-bd)}} .sc-pass{{background:linear-gradient(150deg,#fff 55%,#e2f5e6)}}
.sc-fail::before{{background:var(--fail-bd)}} .sc-fail{{background:linear-gradient(150deg,#fff 55%,#fde9ec)}}
.sc-rate::before{{background:var(--rate-bd)}} .sc-rate{{background:linear-gradient(150deg,#fff 55%,#f0e8fb)}}

.sc-icon{{font-size:22px;margin-bottom:8px}}
.sc-num{{font-size:42px;font-weight:800;line-height:1;margin-bottom:4px;letter-spacing:-1.5px}}
.sc-exec .sc-num{{color:var(--exec-dk)}} .sc-pass .sc-num{{color:var(--pass-dk)}}
.sc-fail .sc-num{{color:var(--fail-dk)}} .sc-rate .sc-num{{color:var(--rate-dk)}}
.sc-lbl{{font-size:10.5px;font-weight:700;text-transform:uppercase;letter-spacing:.8px;color:var(--muted)}}
.sc-bar{{margin-top:12px;height:5px;border-radius:3px;background:#e5e7eb;overflow:hidden}}
.sc-bar-fill{{height:100%;border-radius:3px;transition:width .4s}}
.fill-exec{{background:var(--exec-bd);width:100%}}
.fill-pass{{background:var(--pass-bd)}}
.fill-fail{{background:var(--fail-bd)}}
.fill-rate{{background:var(--rate-bd)}}

/* Donut chart */
.donut-area{{
  width:200px;flex-shrink:0;display:flex;flex-direction:column;
  align-items:center;justify-content:center
}}
.donut-wrap{{position:relative;width:130px;height:130px}}
.donut-center{{
  position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);
  text-align:center;pointer-events:none
}}
.donut-center .dc-num{{font-size:24px;font-weight:800;color:var(--pass-dk);line-height:1}}
.donut-center .dc-sub{{font-size:10px;color:var(--muted);font-weight:600;
  text-transform:uppercase;letter-spacing:.5px}}
.donut-legend{{display:flex;gap:14px;margin-top:10px}}
.dl-item{{display:flex;align-items:center;gap:5px;font-size:11px;color:var(--muted);font-weight:600}}
.dl-dot{{width:9px;height:9px;border-radius:50%;flex-shrink:0}}

/* ══════════════════════════════════════
   FAILURES PANEL
══════════════════════════════════════ */
.fail-panel{{
  background:var(--card);border-radius:12px;
  border:1px solid var(--fail-bd);
  box-shadow:0 2px 8px rgba(183,28,28,.10);
  margin-bottom:24px;overflow:hidden
}}
.fail-panel-hdr{{
  background:linear-gradient(90deg,#9b2a2a 0%,#c62828 100%);
  color:#fff;padding:14px 24px;display:flex;align-items:center;
  justify-content:space-between;cursor:pointer;user-select:none
}}
.fail-panel-hdr h2{{font-size:15px;font-weight:700;display:flex;align-items:center;gap:10px;color:#fff}}
.fail-count-badge{{
  background:rgba(255,255,255,.2);border:1px solid rgba(255,255,255,.35);
  color:#fff;padding:2px 10px;border-radius:10px;font-size:12px
}}
.fail-panel-body{{padding:20px 24px;display:none}}
.fail-panel-body.open{{display:block}}

.failure-group{{margin-bottom:20px}}
.failure-group:last-child{{margin-bottom:0}}
.fg-title{{
  font-size:13px;font-weight:700;color:#fff;
  padding:9px 14px;background:var(--fail-md);
  border-left:4px solid var(--fail-dk);border-radius:0 5px 5px 0;
  margin-bottom:8px;display:flex;align-items:center;gap:10px
}}
.fg-body{{padding-left:4px}}

/* ══════════════════════════════════════
   SEARCH BAR
══════════════════════════════════════ */
.search-wrap{{margin-bottom:20px}}
.search-box{{
  background:var(--card);border:2px solid var(--bdr);border-radius:8px;
  padding:11px 18px;display:flex;align-items:center;gap:12px;
  box-shadow:0 1px 4px rgba(0,0,0,.06);transition:border-color .2s
}}
.search-box:focus-within{{border-color:var(--accent);box-shadow:0 0 0 3px rgba(218,41,28,.12)}}
.search-icon{{color:var(--muted);font-size:15px;flex-shrink:0}}
.search-box input{{
  flex:1;border:none;outline:none;font-size:14px;color:var(--txt);background:transparent
}}
.search-box input::placeholder{{color:#adb5bd}}
.search-hint{{font-size:11px;color:var(--muted);margin-top:5px;padding-left:4px;min-height:16px}}

/* ══════════════════════════════════════
   REPORT CARDS
══════════════════════════════════════ */
.section-heading{{
  font-size:11px;font-weight:700;color:var(--muted);
  text-transform:uppercase;letter-spacing:1.3px;
  margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid var(--bdr)
}}
.report-card{{
  background:var(--card);border-radius:10px;border:1px solid var(--bdr);
  box-shadow:0 1px 6px rgba(0,0,0,.06);margin-bottom:14px;overflow:hidden;
  transition:box-shadow .2s
}}
.report-card:hover{{box-shadow:0 3px 14px rgba(0,0,0,.1)}}

.card-hdr{{
  display:flex;align-items:center;justify-content:space-between;
  padding:13px 20px;cursor:pointer;background:#f6f8fb;
  border-bottom:1px solid var(--bdr);user-select:none
}}
.card-hdr:hover{{background:#edf1f7}}
.card-hdr-left{{display:flex;align-items:center;gap:10px;flex:1;min-width:0}}
.card-icon{{font-size:15px;flex-shrink:0}}
.card-title{{
  font-size:14px;font-weight:700;color:var(--txt);
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis
}}
.sec-badge{{
  font-size:10px;font-weight:700;padding:2px 8px;border-radius:8px;
  flex-shrink:0;text-transform:uppercase;letter-spacing:.3px
}}
.sec-fail{{background:var(--fail-bg);color:var(--fail-dk);border:1px solid var(--fail-bd)}}
.sec-pass{{background:var(--pass-bg);color:var(--pass-dk);border:1px solid var(--pass-bd)}}
.chevron{{color:var(--muted);font-size:12px;transition:transform .25s;flex-shrink:0;margin-left:10px}}
.chevron.open{{transform:rotate(180deg)}}
.card-body{{padding:20px 24px;display:none}}
.card-body.open{{display:block}}

/* Mini stats inside card */
.mini-stats-row{{
  display:flex;gap:22px;margin-bottom:16px;
  padding:10px 14px;background:#f6f8fb;
  border-radius:6px;border:1px solid var(--bdr)
}}
.ms{{font-size:12.5px;font-weight:600}}
.ms-exec{{color:var(--exec-dk)}} .ms-pass{{color:var(--pass-dk)}} .ms-fail{{color:var(--fail-dk)}}

/* ══════════════════════════════════════
   TABLES  (apply to ALL tables in report)
══════════════════════════════════════ */
.tbl-wrap{{width:100%;overflow-x:auto;border-radius:7px;
  border:1px solid var(--bdr);margin:12px 0;
  box-shadow:0 1px 4px rgba(0,0,0,.05)}}
/* min-width ensures scroll kicks in before columns are squeezed */
table{{width:100%;min-width:480px;border-collapse:collapse;font-size:13px;table-layout:auto}}
table th{{
  background:var(--txt);color:#fff;padding:9px 14px;
  text-align:left;font-size:11px;font-weight:700;
  text-transform:uppercase;letter-spacing:.6px;
  white-space:nowrap;min-width:80px
}}
table td{{
  padding:8px 14px;border-bottom:1px solid #f0f1f3;
  vertical-align:middle;line-height:1.45;
  min-width:80px;
  word-break:normal;overflow-wrap:break-word
}}
/* First column (label cells) – never wrap */
table td:first-child{{
  white-space:nowrap;word-break:normal;overflow-wrap:normal
}}
/* Cells classified as "short status" by JS – never wrap (PASS / FAIL / OK / #) */
table td.td-nowrap{{
  white-space:nowrap;word-break:normal;overflow-wrap:normal;
  width:1%;font-weight:700;text-align:center
}}
table tr:last-child td{{border-bottom:none}}
table tr:nth-child(even) td{{background:#fafbfc}}
table tr:hover td{{background:#eef4fb !important;color:var(--txt) !important}}
/* ── Pass / Fail row highlighting ──────────────────────────────────────
   Light tint + left accent border instead of solid opaque fill.
   Dark text throughout so content stays legible on any background. */
table tr.row-fail td{{
  background:#fff0f0 !important;
  border-left:3px solid var(--fail-bd);
  color:var(--txt) !important;font-weight:600
}}
table tr.row-pass td{{
  background:#f0fbf1 !important;
  border-left:3px solid var(--pass-bd);
  color:var(--txt) !important;font-weight:500
}}
table tr.row-fail:hover td{{background:#fde0e0 !important;color:var(--txt) !important}}
table tr.row-pass:hover td{{background:#ddf3df !important;color:var(--txt) !important}}

/* Headings generated by extract functions */
.card-body h2,.fg-body h2,.fail-panel-body h2{{
  font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.9px;
  color:var(--txt);margin:18px 0 8px;padding-bottom:6px;
  border-bottom:2px solid var(--accent);display:flex;align-items:center;gap:8px
}}
.card-body h2:first-child,.fg-body h2:first-child{{margin-top:0}}

/* ══════════════════════════════════════
   BADGES
══════════════════════════════════════ */
.badge{{
  display:inline-block;padding:2px 8px;border-radius:10px;
  font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.4px
}}
.badge-fail{{background:var(--fail-bd);color:#fff}}
.badge-pass{{background:var(--pass-bd);color:#fff}}
.badge-neutral{{background:#e5e7eb;color:var(--muted)}}

/* ══════════════════════════════════════
   PRE / CODE
══════════════════════════════════════ */
pre{{
  background:#1e1e2e;color:#cdd6f4;border-radius:6px;padding:16px;
  font-size:12px;font-family:"Consolas","Courier New",monospace;
  white-space:pre-wrap;word-break:break-all;
  max-height:450px;overflow-y:auto;margin:8px 0;line-height:1.5
}}

/* ══════════════════════════════════════
   IMAGES  (relative-path gallery cards)
══════════════════════════════════════ */
.img-note{{
  font-size:12px;color:var(--accent);background:var(--exec-bg);
  border:1px solid var(--exec-bd);border-radius:6px;
  padding:9px 14px;margin-bottom:14px;display:flex;align-items:center;gap:8px
}}
.img-grid{{
  display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));
  gap:16px;margin-top:4px
}}
.img-figure{{
  border:1px solid var(--bdr);border-radius:8px;overflow:hidden;
  background:#fff;box-shadow:0 1px 5px rgba(0,0,0,.08);
  transition:box-shadow .2s;display:flex;flex-direction:column
}}
.img-figure:hover{{box-shadow:0 5px 18px rgba(0,0,0,.16)}}
.img-thumb-link{{
  display:block;overflow:hidden;background:#f6f8fb;height:180px;flex-shrink:0
}}
.img-figure img{{
  width:100%;height:180px;object-fit:contain;display:block;
  transition:transform .3s
}}
.img-figure:hover img{{transform:scale(1.04)}}
.img-figcap{{
  padding:8px 10px;border-top:1px solid var(--bdr);
  display:flex;flex-direction:column;gap:3px
}}
.img-name{{
  font-size:11.5px;color:var(--txt);font-weight:600;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;display:block
}}
.img-open-link{{
  font-size:10.5px;color:var(--accent);text-decoration:none;
  display:inline-flex;align-items:center;gap:3px
}}
.img-open-link:hover{{text-decoration:underline}}

/* ══════════════════════════════════════
   MISC HELPERS
══════════════════════════════════════ */
.no-failures{{
  color:var(--pass-dk);background:var(--pass-bg);border:1px solid var(--pass-bd);
  border-radius:6px;padding:12px 18px;font-size:13px;font-weight:600;
  display:flex;align-items:center;gap:8px
}}
.muted-note{{color:var(--muted);font-style:italic;font-size:13px;padding:8px 0}}
.file-list{{list-style:none}}
.file-list li{{
  padding:6px 0;border-bottom:1px solid #f5f5f5;font-size:13px;color:var(--muted);
  display:flex;align-items:center;gap:8px
}}
.file-list li:last-child{{border-bottom:none}}
.hidden{{display:none !important}}
/* ── Report link note (filename reference under tables) ── */
.report-link-note{{
  font-size:11.5px;color:var(--muted);font-style:italic;
  display:inline-flex;align-items:center;gap:4px
}}

/* ══════════════════════════════════════
   EMBEDDED HTML REPORT IFRAMES
   Individual HTML reports are rendered in
   sandboxed iframes so their original styling
   and interactivity are fully preserved.
══════════════════════════════════════ */
.embedded-report-title{{
  font-size:13px;font-weight:700;color:var(--txt);
  margin:18px 0 6px;display:flex;align-items:center;gap:6px
}}
.embedded-report-title:first-child{{margin-top:0}}
.embedded-report-wrap{{
  border:1px solid var(--bdr);border-radius:8px;overflow:hidden;
  box-shadow:0 2px 8px rgba(0,0,0,.08);margin:0 0 16px;background:#fff
}}
.embedded-report-frame{{
  width:100%;height:640px;border:none;display:block;background:#fff
}}

/* ══════════════════════════════════════
   FOOTER
══════════════════════════════════════ */
.report-footer{{
  text-align:center;padding:20px;font-size:11px;color:var(--muted);
  border-top:1px solid var(--bdr);margin-top:24px
}}

/* ══════════════════════════════════════
   PRINT
══════════════════════════════════════ */
@media print{{
  .sidebar,.search-wrap,.page-hdr{{display:none!important}}
  .main-wrap{{margin-left:0!important}}
  .card-body,.fail-panel-body{{display:block!important}}
  .stats-dash,.fail-panel,.report-card{{break-inside:avoid}}
}}
</style>
<script>
/* ── CARD TOGGLE ── */
function toggleCard(id) {{
  var body = document.getElementById('body-' + id);
  var chv  = document.getElementById('chv-'  + id);
  if (!body) return;
  var open = body.classList.toggle('open');
  if (chv) chv.classList.toggle('open', open);
}}

/* ── SHOW + SCROLL TO SECTION ── */
function showSection(id) {{
  var body = document.getElementById('body-' + id);
  var chv  = document.getElementById('chv-'  + id);
  if (!body) return;
  body.classList.add('open');
  if (chv) chv.classList.add('open');
  var card = document.getElementById(id);
  if (card) card.scrollIntoView({{behavior:'smooth',block:'start'}});
}}

/* ── FAILURES PANEL TOGGLE ── */
function toggleFailPanel() {{
  var body = document.getElementById('fail-panel-body');
  if (!body) return;
  body.classList.toggle('open');
}}

/* ── LIVE SEARCH ── */
function searchReports() {{
  var q     = document.getElementById('searchInput').value.toLowerCase().trim();
  var cards = document.querySelectorAll('.report-card');
  var vis   = 0;
  cards.forEach(function(card) {{
    if (!q) {{
      card.classList.remove('hidden');
      vis++;
      return;
    }}
    if (card.textContent.toLowerCase().includes(q)) {{
      card.classList.remove('hidden');
      vis++;
      var body = card.querySelector('.card-body');
      var chv  = card.querySelector('.chevron');
      if (body) body.classList.add('open');
      if (chv)  chv.classList.add('open');
    }} else {{
      card.classList.add('hidden');
    }}
  }});
  var hint = document.getElementById('searchHint');
  if (hint) {{
    hint.textContent = q
      ? (vis + ' of ' + cards.length + ' report(s) match \u201c' + q + '\u201d')
      : '';
  }}
}}

/* ── HIGHLIGHT FAIL / PASS ROWS ── */
function styleTableRows() {{
  /* Status keywords that must never wrap */
  var STATUS_RE = new RegExp('^(fail|pass|ok|error|skip|n/a|yes|no|[0-9]{1,5})$', 'i');
  document.querySelectorAll('table tbody tr').forEach(function(row) {{
    var t = row.textContent.toLowerCase();
    if (t.includes('fail')) {{
      row.classList.add('row-fail');
    }} else if (t.includes('pass') || t.includes(' ok') || t.includes('success')) {{
      row.classList.add('row-pass');
    }}
    /* Tag short status-value cells so they never word-wrap */
    row.querySelectorAll('td').forEach(function(td) {{
      var val = td.textContent.trim();
      if (STATUS_RE.test(val)) {{
        td.classList.add('td-nowrap');
      }}
    }});
  }});
}}

/* ── WRAP TABLES IN SCROLLABLE DIV ── */
function wrapTables() {{
  document.querySelectorAll(
    '.card-body table, .fg-body table, .fail-panel-body table'
  ).forEach(function(tbl) {{
    if (tbl.parentNode.classList.contains('tbl-wrap')) return;
    var w = document.createElement('div');
    w.className = 'tbl-wrap';
    tbl.parentNode.insertBefore(w, tbl);
    w.appendChild(tbl);
  }});
}}

window.addEventListener('DOMContentLoaded', function() {{
  styleTableRows();
  wrapTables();
}});
</script>
</head>
<body>

<!-- ══════════════ SIDEBAR ══════════════ -->
<nav class="sidebar">
  <div class="sb-brand">
    <div class="brand-icon">
      <img src="data:image/png;base64,{_MAGNA_LOGO_B64}" alt="Magna Electronics" style="display:block;width:100%;height:auto">
    </div>
    <h2>Magna Electronics</h2>
    <div class="brand-sub">ADAS &bull; Surround View Camera System</div>
  </div>

  <div class="sb-kpi">
    <div class="kpi"><div class="kpi-num kn-exec">{total_executed}</div><div class="kpi-lbl">Executed</div></div>
    <div class="kpi"><div class="kpi-num kn-pass">{total_passed}</div><div class="kpi-lbl">Passed</div></div>
    <div class="kpi"><div class="kpi-num kn-fail">{total_failed}</div><div class="kpi-lbl">Failed</div></div>
  </div>

  <div class="nav-section-lbl">Quick Jump</div>
  <ul class="nav-list">
    <li class="nav-item">
      <a class="nav-link" href="javascript:void(0);" onclick="toggleFailPanel()">
        <span class="nav-title">&#x26A0; Failure Summary</span>
        <span class="nav-badge nb-fail">{total_failed}</span>
      </a>
    </li>
    <li class="nav-item">
      <a class="nav-link" href="javascript:void(0);" onclick="showSection('logs')">
        <span class="nav-title">&#x1F4C4; Logs &amp; Info</span>
      </a>
    </li>
    <li class="nav-item">
      <a class="nav-link" href="javascript:void(0);" onclick="showSection('images')">
        <span class="nav-title">&#x1F5BC; Embedded Images</span>
      </a>
    </li>
  </ul>

  <div class="nav-section-lbl">Embedded Reports ({num_reports})</div>
  <ul class="nav-list">
    {navigation_links}
  </ul>
</nav>

<!-- ══════════════ MAIN CONTENT ══════════════ -->
<div class="main-wrap">

  <header class="page-hdr">
    <div>
      <h1>
        <img src="data:image/png;base64,{_MAGNA_LOGO_SM_B64}" width="36" height="36" alt="Magna Electronics" style="flex-shrink:0;display:block">
        Magna Electronics &ndash; Software Test Report
      </h1>
      <div class="hdr-sub">
        ADAS &bull; Surround View Camera System &nbsp;&bull;&nbsp;
        Jenkins CI/CD Pipeline &nbsp;&bull;&nbsp;
        {num_reports} embedded report(s) &nbsp;&bull;&nbsp;
        Generated: {report_timestamp}
      </div>
    </div>
    <div class="hdr-pill">&#x1F552; {report_timestamp}</div>
  </header>

  <div class="content">

    <!-- ══ STATISTICS DASHBOARD ══ -->
    <div class="stats-dash">
      <div class="dash-title">&#x1F4C8; Executive Test Summary &mdash; ADAS Surround View Camera &mdash; All Reports Combined</div>
      <div class="stats-row">
        <div class="stat-card sc-exec">
          <div class="sc-icon">&#x1F3AF;</div>
          <div class="sc-num">{total_executed}</div>
          <div class="sc-lbl">Total Executed</div>
          <div class="sc-bar"><div class="sc-bar-fill fill-exec"></div></div>
        </div>
        <div class="stat-card sc-pass">
          <div class="sc-icon">&#x2705;</div>
          <div class="sc-num">{total_passed}</div>
          <div class="sc-lbl">Tests Passed</div>
          <div class="sc-bar"><div class="sc-bar-fill fill-pass" style="width:{pass_rate}%"></div></div>
        </div>
        <div class="stat-card sc-fail">
          <div class="sc-icon">&#x274C;</div>
          <div class="sc-num">{total_failed}</div>
          <div class="sc-lbl">Tests Failed</div>
          <div class="sc-bar"><div class="sc-bar-fill fill-fail" style="width:{fail_rate}%"></div></div>
        </div>
        <div class="stat-card sc-rate">
          <div class="sc-icon">{pass_rate_icon}</div>
          <div class="sc-num">{pass_rate}%</div>
          <div class="sc-lbl">Pass Rate</div>
          <div class="sc-bar"><div class="sc-bar-fill fill-rate" style="width:{pass_rate}%"></div></div>
        </div>
        <!-- SVG Donut Chart (self-contained, no CDN) -->
        <div class="donut-area">
          <div class="donut-wrap">
            <svg viewBox="0 0 130 130" width="130" height="130" aria-label="Pass/Fail donut chart">
              <!-- background ring -->
              <circle cx="65" cy="65" r="54" fill="none" stroke="#e5e7eb" stroke-width="16"/>
              <!-- pass arc -->
              <circle cx="65" cy="65" r="54" fill="none" stroke="#43a047" stroke-width="16"
                stroke-dasharray="{pass_arc} {circ}"
                stroke-dashoffset="0"
                transform="rotate(-90 65 65)"/>
              <!-- fail arc (offset starts right after pass arc) -->
              <circle cx="65" cy="65" r="54" fill="none" stroke="#ef5350" stroke-width="16"
                stroke-dasharray="{fail_arc} {circ}"
                stroke-dashoffset="{fail_arc_offset}"
                transform="rotate(-90 65 65)"/>
            </svg>
            <div class="donut-center">
              <div class="dc-num">{pass_rate}%</div>
              <div class="dc-sub">Pass Rate</div>
            </div>
          </div>
          <div class="donut-legend">
            <div class="dl-item"><div class="dl-dot" style="background:#43a047"></div>Pass</div>
            <div class="dl-item"><div class="dl-dot" style="background:#ef5350"></div>Fail</div>
          </div>
        </div>
      </div>
    </div>

    <!-- ══ CONSOLIDATED FAILURES PANEL ══ -->
    <div class="fail-panel">
      <div class="fail-panel-hdr" onclick="toggleFailPanel()">
        <h2>
          &#x26A0;&#xFE0F; Consolidated Failure Summary
          <span class="fail-count-badge">{total_failed} Failure(s) Across All Reports</span>
        </h2>
        <span style="font-size:12px;opacity:.7">&#x25BC; expand / collapse</span>
      </div>
      <div class="fail-panel-body" id="fail-panel-body">
        {consolidated_failures}
      </div>
    </div>

    <!-- ══ SEARCH ══ -->
    <div class="search-wrap">
      <div class="search-box">
        <span class="search-icon">&#x1F50D;</span>
        <input type="text" id="searchInput"
          placeholder="Search reports by name, test case, keyword, status&hellip;"
          oninput="searchReports()">
      </div>
      <div class="search-hint" id="searchHint"></div>
    </div>

    <!-- ══ INDIVIDUAL REPORT CARDS ══ -->
    <div class="section-heading">&#x1F4CB; Embedded Reports ({num_reports} total)</div>
    {report_sections_html}

    <!-- ══ LOGS ══ -->
    <div class="report-card" id="logs">
      <div class="card-hdr" onclick="toggleCard('logs')">
        <div class="card-hdr-left">
          <span class="card-icon">&#x1F4C4;</span>
          <span class="card-title">Logs &amp; Additional Information</span>
        </div>
        <span class="chevron" id="chv-logs">&#9660;</span>
      </div>
      <div class="card-body" id="body-logs">
        {text_content if text_content else '<p class="muted-note">No logs available.</p>'}
      </div>
    </div>

    <!-- ══ IMAGES ══ -->
    <div class="report-card" id="images">
      <div class="card-hdr" onclick="toggleCard('images')">
        <div class="card-hdr-left">
          <span class="card-icon">&#x1F5BC;</span>
          <span class="card-title">Embedded Images</span>
        </div>
        <span class="chevron" id="chv-images">&#9660;</span>
      </div>
      <div class="card-body" id="body-images">
        <p class="img-note">
          &#x2139; Images are stored in the <strong>report_images/</strong> subfolder alongside this file.
          <strong>Do not delete or move the report_images/ folder</strong> — doing so will break all image links.
          When sharing this report, always include the report_images/ folder.
          Click any thumbnail to open at full resolution in a new tab.
        </p>
        <div class="img-grid">
          {image_tags if image_tags else '<p class="muted-note">No images found.</p>'}
        </div>
      </div>
    </div>

    <!-- ══ COPIED FILES ══ -->
    <div class="report-card" id="copied-files">
      <div class="card-hdr" onclick="toggleCard('copied-files')">
        <div class="card-hdr-left">
          <span class="card-icon">&#x1F4C1;</span>
          <span class="card-title">Copied Files</span>
        </div>
        <span class="chevron" id="chv-copied-files">&#9660;</span>
      </div>
      <div class="card-body" id="body-copied-files">
        {files_list_html}
      </div>
    </div>

    <div class="report-footer">
      Magna Electronics &bull; ADAS Surround View Camera System &bull;
      Software Test Report &bull; Jenkins CI/CD &bull; {report_timestamp}
      &bull; {num_reports} report(s) embedded
    </div>

  </div><!-- /content -->
</div><!-- /main-wrap -->
</body>
</html>"""

        output_path = os.path.join(destination_path, "Software_Test_Report.html")
        with open(output_path, "w", encoding="utf-8") as html_out:
            html_out.write(html_template)

        logging.info(f"Report successfully generated at {output_path}")

    except Exception as e:
        logging.critical(f"Unexpected error generating the report: {e}")

def extract_results_table(soup):
    results_table = []
    start_write = False

    for div in soup.find_all('div'):
        # Start 
        if "Results" in div.get_text():
            start_write = True

        # Stop 
        if "Test Configuration Information" in div.get_text() and start_write:
            start_write = False

        # look for the table containing the results
        if start_write and div.find('table'):
            table = div.find('table')
            for row in table.find_all('tr'):
                columns = [col.get_text(strip=True) for col in row.find_all('td')]
                
                # start adding rows once the first column is "1"
                if len(columns) > 0 and columns[0] == "1":
                    if len(columns) >= 3 and columns[2].lower() == "fail": 
                        results_table.append(columns[:3]) 

                # keep capturing only 'fail' rows
                elif len(results_table) > 0 and len(columns) >= 3 and columns[2].lower() == "fail":
                    results_table.append(columns[:3])  
    
    return results_table
def extract_keyword_table(soup, keyword):
    """Extracts tables related to the specific keywords from the HTML report."""
    keyword_tables = {keyword: [] for keyword in keywords}
    found_keyword_section = {keyword: False for keyword in keywords}
    parsed_tables = {keyword: False for keyword in keywords} 

    for table in soup.find_all('table'):
        rows = table.find_all('tr')
        for row in rows:
            text = row.get_text(strip=True)

            for keyword in keywords:
                # check both conditions 
                if (f"CAN signal '{keyword}'" in text or f"_22{keyword}" in text) and not parsed_tables[keyword]:  
                    found_keyword_section[keyword] = True

                if found_keyword_section[keyword] and "Check of expected values" in text:
                    # Look for the next table
                    response_table = row.find_next('table', class_='InfoTableExpand')
                    if response_table:
                        # Extract data from the found table
                        for data_row in response_table.find_all('tr'):
                            columns = [col.get_text(strip=True) for col in data_row.find_all('td')]
                            if columns:  
                                keyword_tables[keyword].append(columns)
                        # Mark this keyword's table as parsed
                        parsed_tables[keyword] = True
                    # Stop after the first match for this keyword
                    found_keyword_section[keyword] = False

    return keyword_tables


def _cell_plain_text(cell):
    """Return plain text of a BeautifulSoup cell tag, stripping any nested
    <table> elements so sub-table content doesn't bleed into the output column."""
    clone = copy.deepcopy(cell)
    for nested in clone.find_all("table"):
        nested.decompose()
    return clone.get_text(separator=" ", strip=True)


def extract_statistics_table(soup):
    """Extracts the OverviewTable and renders it as a clean, flat two-column table.

    The raw source HTML is parsed and re-rendered so that any inline styles,
    nested elements, or class names from the original test tool are stripped.
    This prevents distorted column widths in the consolidated report.
    """
    statistics_table = soup.find("table", class_="OverviewTable")
    if not statistics_table:
        return "<p><b>No statistics table found in the report.</b></p>"

    rows_html = ""
    for row in statistics_table.find_all("tr"):
        # Only process rows that belong directly to the OverviewTable itself
        if row.find_parent("table") is not statistics_table:
            continue
        cells = row.find_all(["th", "td"], recursive=False)
        if not cells:
            continue
        is_header = bool(row.find("th", recursive=False))
        cell_texts = [html_escape_lib.escape(_cell_plain_text(c)) for c in cells]
        if is_header:
            rows_html += "<tr>" + "".join(f"<th>{t}</th>" for t in cell_texts) + "</tr>"
        else:
            rows_html += "<tr>" + "".join(f"<td>{t}</td>" for t in cell_texts) + "</tr>"

    if not rows_html:
        return "<p><b>No statistics data found in the report.</b></p>"

    return (
        f'<h2>Test Statistics</h2>'
        f'<div class="tbl-wrap">'
        f'<table><tbody>{rows_html}</tbody></table>'
        f'</div>'
    )

def extract_failed_tests(soup, source_filename=None):
    """Extracts a high-level summary of failed test cases.

    Only genuine numbered test-case rows are included (first cell must match
    the step-number pattern, e.g. "2.1.1.2.1").  OverviewTable summary rows,
    section-header rows, timestamp rows, and expand-button rows are all
    excluded so the output is a clean three-column table:
      Step Number | Test Case Name | Result

    Nested sub-tables are never emitted — the output is always a flat table.
    If *source_filename* is provided a note is appended inviting the reader
    to open the original file for the full drill-down detail.
    """
    import re as _re
    # Pattern for a genuine test-case step number: at least two dot-separated
    # numeric segments, e.g. "2.1", "2.1.1.2.1".  Rejects plain integers ("2"),
    # timestamps ("2608.730519"), and button labels ("[-]").
    STEP_RE = _re.compile(r'^\d+(\.\d+){2,}$')

    failed_rows = []
    seen: set[tuple] = set()

    for table in soup.find_all("table"):
        # Skip tables nested inside a table cell (sub-tables)
        if table.find_parent("td") or table.find_parent("th"):
            continue
        # Skip the OverviewTable — it contains summary sentences, not test cases
        if "OverviewTable" in (table.get("class") or []):
            continue

        for row in table.find_all("tr"):
            # Only direct rows of *this* top-level table
            if row.find_parent("table") is not table:
                continue
            # Skip header rows
            if row.find("th", recursive=False):
                continue
            cells = row.find_all("td", recursive=False)
            if not cells:
                continue

            # First-cell text: must match the step-number pattern
            first = _cell_plain_text(cells[0])
            if not STEP_RE.match(first):
                continue

            # Extract plain text for all direct cells (strip nested tables)
            cell_texts = [_cell_plain_text(c) for c in cells]

            # The row must contain "fail" somewhere in its cells
            if not any("fail" in t.lower() for t in cell_texts):
                continue

            row_key = tuple(cell_texts[:3])
            if row_key in seen:
                continue
            seen.add(row_key)

            # Emit exactly 3 columns: step number | test case name | result
            failed_rows.append(cell_texts[:3])

    if not failed_rows:
        return ""

    # Optional reference note pointing to the original source file
    link_html = ""
    if source_filename:
        safe_name = html_escape_lib.escape(os.path.basename(source_filename))
        link_html = (
            f'<span class="report-link-note">'
            f'&#x1F4C4; Full details: <strong>{safe_name}</strong>'
            f' &mdash; see <em>Logs &amp; Info</em> section'
            f'</span>'
        )

    rows_html = "".join(
        "<tr>" + "".join(f"<td>{html_escape_lib.escape(t)}</td>" for t in row_cells) + "</tr>"
        for row_cells in failed_rows
    )

    header_row = "<tr><th>Step</th><th>Test Case</th><th>Result</th></tr>"

    return (
        f'<h2>Failed Tests Summary{link_html}</h2>'
        f'<div class="tbl-wrap">'
        f'<table><thead>{header_row}</thead><tbody>{rows_html}</tbody></table>'
        f'</div>'
    )

def extract_keyword_from_tables(soup, keyword, source_filename=None):
    """Extracts a clean summary of table rows containing the given keyword.

    Only rows that belong directly to top-level tables are searched; nested
    sub-tables inside table cells and their rows are skipped to prevent
    tables-within-tables in the output.  Cell content is rendered as plain
    text with nested markup removed, and trimmed to at most four columns.

    If *source_filename* is provided a note is appended inviting the reader
    to open the original file for detailed drill-down information.
    """
    keyword_lower = keyword.lower()
    extracted_rows = []
    seen: set[tuple] = set()

    for table in soup.find_all("table"):
        # Skip tables nested inside a table cell (sub-tables)
        if table.find_parent("td") or table.find_parent("th"):
            continue

        for row in table.find_all("tr"):
            # Only process rows that belong directly to *this* top-level table
            if row.find_parent("table") is not table:
                continue
            # Only direct-child cells (recursive=False prevents picking up
            # td/th elements that live inside nested sub-tables)
            cells = row.find_all(["th", "td"], recursive=False)
            if not cells:
                continue
            # Deepcopy each cell and strip nested tables before extracting text
            cell_texts = [_cell_plain_text(c) for c in cells]
            row_key = tuple(cell_texts[:3])
            if row_key in seen:
                continue
            if any(keyword_lower in t.lower() for t in cell_texts):
                seen.add(row_key)
                extracted_rows.append(cell_texts[:4])

    if not extracted_rows:
        return ""

    rows_html = "".join(
        "<tr>" + "".join(f"<td>{html_escape_lib.escape(t)}</td>" for t in row_cells) + "</tr>"
        for row_cells in extracted_rows
    )

    link_note = ""
    if source_filename:
        safe_name = html_escape_lib.escape(os.path.basename(source_filename))
        link_note = (
            f'<p class="report-link-note">'
            f'&#x1F4C4; Full details: <strong>{safe_name}</strong>'
            f' &mdash; see <em>Logs &amp; Info</em> section'
            f'</p>'
        )

    return (
        f'<div class="tbl-wrap">'
        f'<table><tbody>{rows_html}</tbody></table>'
        f'</div>'
        f'{link_note}'
    )

def extract_overview_table(soup):
    """Extracts executed, pass, and fail count from the 'OverviewTable'."""
    global pass_count, fail_count, executed_count

    overview_table = soup.find("table", class_="Statistics")
    if not overview_table:
        return "<p><b>OverviewTable not found in the report.</b></p>"

    rows = overview_table.find_all("tr")
    extracted_table = "<h2>Test Summary</h2><table border='1'>"

    for row in rows:
        cells = row.find_all("td")
        row_text = [cell.get_text(strip=True) for cell in cells]

        if "Executed test cases" in row_text[0]:
            executed_count = int(row_text[1])
        elif "Test cases passed" in row_text[0]:
            pass_count = int(row_text[1])
        elif "Test cases failed" in row_text[0]:
            fail_count = int(row_text[1])

        extracted_table += f"<tr>{''.join(f'<td>{cell}</td>' for cell in row_text)}</tr>"

    extracted_table += "</table>"
    return extracted_table

# Allowed file extensions (only images & text files)
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "txt", "html"}

def detect_encoding(file_path):
    """Detects file encoding before reading."""
    with open(file_path, "rb") as f:
        raw_data = f.read(4096)  # Read the first 4KB for detection
        result = chardet.detect(raw_data)
        return result["encoding"] if result["encoding"] else "utf-8"

def get_unique_filename(dest_dir, filename):
    """Ensures a unique filename by appending a number if it already exists."""
    base, ext = os.path.splitext(filename)
    counter = 1
    new_filename = filename

    while os.path.exists(os.path.join(dest_dir, new_filename)):
        new_filename = f"{base}_{counter}{ext}"
        counter += 1

    return new_filename

def get_file_hash(file_path, chunk_size=65536):
    """Generate an MD5 hash for a file using chunked reads to avoid memory spikes
    on large images or log files."""
    hasher = hashlib.md5()
    with open(file_path, "rb") as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            hasher.update(chunk)
    return hasher.hexdigest()

def copy_and_embed_files(source_paths, destination_path, delete_after_embedding=True):
    """Copies and embeds text/HTML files inline and images as relative-path gallery cards.

    Images are placed in a flat 'report_images/' subfolder next to the HTML output so
    that the report file itself stays small.  Text and HTML files are still embedded
    inline so the report is self-contained for log/trace data.  Only embedded text
    files are deleted after processing; image files are intentionally kept because the
    HTML references them by relative path.
    """

    image_tags = ""
    text_content = ""
    image_count = 0
    text_count = 0
    copied_files = []          # All successfully copied files (for the "Copied Files" section)
    text_files_to_delete = []  # Only text/HTML files are deleted after inline embedding
    embedded_hashes = set()    # Deduplicate by content hash

    if not os.path.exists(destination_path):
        os.makedirs(destination_path)

    # Flat subfolder that sits next to the HTML output file
    images_dir = os.path.join(destination_path, "report_images")

    for source_path in source_paths:
        if not os.path.exists(source_path):
            logging.warning(f"Skipping non-existent path: {source_path}")
            continue

        for root, _, files in os.walk(source_path):
            for filename in files:
                file_ext = filename.lower().split(".")[-1]

                if file_ext not in ALLOWED_EXTENSIONS:
                    logging.info(f"Skipping unsupported file: {filename}")
                    continue

                src_file = os.path.join(root, filename)

                try:
                    if file_ext in ["png", "jpg", "jpeg", "gif"]:
                        # ── Images: copy to flat report_images/ subfolder ──────────
                        # Hash the source first to skip copying duplicate content
                        src_hash = get_file_hash(src_file)
                        if src_hash in embedded_hashes:
                            logging.info(f"Skipping duplicate image (source): {filename}")
                            continue

                        os.makedirs(images_dir, exist_ok=True)
                        unique_filename = get_unique_filename(images_dir, filename)
                        dest_file = os.path.join(images_dir, unique_filename)
                        shutil.copy2(src_file, dest_file)

                        if not os.path.exists(dest_file):
                            continue

                        embedded_hashes.add(src_hash)
                        copied_files.append(dest_file)

                        # Relative URL path (forward slashes for HTML on all OS).
                        # Escape the filename for safe use in HTML attributes and text.
                        safe_name = html_escape_lib.escape(unique_filename, quote=True)
                        rel_url = "report_images/" + unique_filename
                        image_tags += (
                            f'<figure class="img-figure">'
                            f'<a class="img-thumb-link" href="{rel_url}" target="_blank"'
                            f'   rel="noopener noreferrer"'
                            f'   title="Click to open full size: {safe_name}">'
                            f'<img src="{rel_url}" alt="{safe_name}" loading="lazy">'
                            f'</a>'
                            f'<div class="img-figcap">'
                            f'<span class="img-name" title="{safe_name}">{safe_name}</span>'
                            f'<a class="img-open-link" href="{rel_url}" target="_blank"'
                            f'   rel="noopener noreferrer">'
                            f'&#x1F517; Open full size</a>'
                            f'</div>'
                            f'</figure>\n'
                        )
                        image_count += 1
                        # NOTE: image files are NOT added to text_files_to_delete –
                        # they must remain on disk so the HTML links resolve correctly.

                    elif file_ext in ["txt", "html"]:
                        # ── Text/HTML: preserve folder structure, embed inline ─────
                        # Hash the source first to skip copying duplicate content
                        src_hash = get_file_hash(src_file)
                        if src_hash in embedded_hashes:
                            logging.info(f"Skipping duplicate text file (source): {filename}")
                            continue

                        relative_path = os.path.relpath(root, source_path)
                        dest_dir = os.path.join(destination_path, relative_path)
                        os.makedirs(dest_dir, exist_ok=True)
                        unique_filename = get_unique_filename(dest_dir, filename)
                        dest_file = os.path.join(dest_dir, unique_filename)
                        shutil.copy2(src_file, dest_file)

                        if not os.path.exists(dest_file):
                            continue

                        embedded_hashes.add(src_hash)
                        copied_files.append(dest_file)

                        text_count += 1
                        if file_ext == "html":
                            # ── HTML reports: render in an iframe so their original
                            #    styling and interactivity are fully preserved.
                            #    The file stays on disk; only the <src> path is embedded.
                            rel_url = os.path.relpath(dest_file, destination_path).replace("\\", "/")
                            safe_title = html_escape_lib.escape(unique_filename, quote=True)
                            text_content += (
                                f'<h3 class="embedded-report-title">'
                                f'{safe_title}</h3>'
                                f'<div class="embedded-report-wrap">'
                                f'<iframe src="{rel_url}" class="embedded-report-frame"'
                                f' title="{safe_title}" loading="lazy"></iframe>'
                                f'</div>\n'
                            )
                            # Do NOT mark for deletion – the iframe src must stay on disk.
                        else:
                            # ── Plain text: HTML-escape content so angle brackets and
                            #    special chars are displayed literally, not rendered.
                            encoding = detect_encoding(dest_file)
                            with open(dest_file, "r", encoding=encoding, errors="replace") as txt_file:
                                text_content += (
                                    f'<h3>{html_escape_lib.escape(unique_filename)}</h3>'
                                    f'<pre>{html_escape_lib.escape(txt_file.read())}</pre><br>'
                                )
                            text_files_to_delete.append(dest_file)

                except Exception as e:
                    logging.error(f"Failed to process {src_file}: {e}")

    logging.info(f"Total text/HTML files embedded: {text_count}")
    logging.info(f"Total image files linked: {image_count}")
    logging.info(f"Total successfully copied files: {len(copied_files)}")

    # Delete only the inline-embedded text/HTML files (images stay for the links)
    if delete_after_embedding:
        for file_path in text_files_to_delete:
            try:
                os.remove(file_path)
                logging.info(f"Deleted embedded text file: {file_path}")
            except Exception as e:
                logging.error(f"Failed to delete {file_path}: {e}")

    return text_content, image_tags, copied_files

def getCPULoadResults():
    script_name = "CPU_Load_Graph_Parser.py"
    script_path = None

    # Search for the script in DEFAULT_SOURCE_PATHS
    for path in DEFAULT_SOURCE_PATHS:
        candidate_path = os.path.join(path, script_name)
        if os.path.isfile(candidate_path):
            script_path = candidate_path
            break

    if not script_path:
        print(f"Error: {script_name} not found in DEFAULT_SOURCE_PATHS!")
        return

    print(f"Found {script_name} at: {script_path}")

    # Extract directory where the script is located
    script_dir = os.path.dirname(script_path)

    command_args = ["python", script_path]
    try:
        result = subprocess.run(command_args, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, cwd=script_dir)
        
        print(result.stdout)
        if result.returncode != 0:
            print(f"Error occurred:\n{result.stderr}")
        else:
            print("CPU Load processing complete. Now cleaning workspaces...")

    except Exception as e:
        print(f"Failed to start {script_name}: {e}")

if __name__ == '__main__':
    
    # Retrieve Jenkins workspace environment variable
    jenkins_workspace = os.environ.get('WORKSPACE')

    if jenkins_workspace:
        base_path = os.path.join(jenkins_workspace, "BVTRBS", "03_VariantDependent", "Customer")
    
        # Find the folder dynamically using glob
        matching_folders = glob.glob(os.path.join(base_path, "*"))
    
        if matching_folders:
            dynamic_path = matching_folders[0]  # Pick the first matching folder
        else:
            print("Warning: No matching folder found. Falling back to base path.")
            dynamic_path = base_path  # Fallback to base path if no folder is found
    else:
        print("Warning: WORKSPACE environment variable not set. Falling back to the current working directory.")
        dynamic_path = os.getcwd()  # Use the current working directory as the fallback path

    # Add dynamic_path to default source paths
    DEFAULT_SOURCE_PATHS.append(dynamic_path)

    print(f"Using dynamic path: {dynamic_path}")  # Debugging print

    parser = argparse.ArgumentParser(description="Generate an HTML report with extracted test data, failures, logs, and images.")
    parser.add_argument('html_file', type=str, nargs='?', default="Report_Sanity.html", help="Path to the HTML file. If not provided, the script will find one automatically.")
    parser.add_argument('keyword', type=str, nargs='?', default="Evaluate response", help="Keyword to search for in the file.")
    parser.add_argument('--source_paths', type=str, nargs='+', default=DEFAULT_SOURCE_PATHS, help="List of source directories.")
    parser.add_argument('--destination_path', type=str, default=DEFAULT_DESTINATION_PATH, help="Destination path for copied files.")
    args = parser.parse_args()

    #Check for cpu load results if found
    getCPULoadResults()

    #Automatically search for all report files in source paths
    found_html_files = find_html_files(DEFAULT_SOURCE_PATHS)
    if not found_html_files:
        exit(1)

    #Generate a consolidated Software Test Report of all HTML files found
    generate_html_report(found_html_files, args.source_paths, args.destination_path, args.keyword)

    #Clean up work spaces of old report files for next iterations on jenkins node
    print(f"Cleaning Workspaces...")
    os.system("del /f /q /s C:\\JS\\ws\\develop\\sw\\Release\\FlashLog.txt")
    os.system("del /f /q /s C:\\JS\\CT_Server_Standard_files_Continous_Testing\\python\\*.html")
    os.system("del /f /q /s C:\\JS\\CT_Server_Standard_files_Continous_Testing\\python\\*.txt")
    os.system("del /f /q /s C:\\JS\\CT_Server_Standard_files_Continous_Testing\\python\\*.xlsx")
    os.system("del /f /q /s C:\\JS\\CT_Server_Standard_files_Continous_Testing\\python\\*jpg")
   
